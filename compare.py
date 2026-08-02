#!/usr/bin/env python3
"""
Compare algorithm-generated mystery boxes against historical manual allocations.

For each offer, computes per-box metrics for both manual and algorithm allocations:
  - Total value (cents) and value as % of target
  - Unique item count
  - Fruit/veg value split (% fruit)
  - Fungible group duplicates (same fungible group appearing multiple times)
  - Preference compliance (fruit-only/veg-only violations)

Prints a summary comparison table with averages across all boxes.
"""

import csv
import json
import logging
import os
import sys
import time
from collections import defaultdict
from pathlib import Path


# Suppress INFO/DEBUG noise from allocator library and paramiko during batch comparison
logging.getLogger("paramiko").setLevel(logging.WARNING)
logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(name)s: %(message)s")

from allocator.allocator import allocate, build_boxes_from_db
from allocator.box_parser import infer_box_tier
from allocator.categorizer import assign_classification, assign_fungible_group
from allocator.config import (
    BOX_TIERS,
    CATEGORY_FRUIT,
    CATEGORY_VEGETABLES,
    DIVERSITY_PENALTY_MULTIPLIER,
    DIVERSITY_WEIGHTS,
    DIVERSITY_FALLBACK_SCORE,
    DONATION_IDENTIFIERS,
    FUNGIBLE_GROUPS,
    GROUP_ALLOWANCES,
    GROUP_CONCENTRATION_MULTIPLIER,
    GROUP_QTY_EXPONENT,
    MAX_COMPOSITE_SCORE,
    MAX_VALUE_SHARE_MULTIPLIER,
    MAX_VALUE_SHARE_THRESHOLD,
    PACK_PRICE_TOLERANCE_CENTS,
    PREF_VIOLATION_PENALTY,
    QUANTITY_CLASSES,
    SAME_ITEM_MULTIPLIER,
    SIZE_FLOOR_MULTIPLIER,
    SIZE_FLOOR_TARGETS,
    SLOT_DEGREE_THRESHOLD,
    detect_pack_size,
)
from allocator.strategies._scoring import (
    _resolve_item_allowance_from_lookup,
    value_penalty,
)
from allocator.strategies import DEFAULT_STRATEGY, list_strategies
from allocator.db import fetch_categories, fetch_mystery_box_buyers, fetch_offer_items

CLEANED_DIR = Path(__file__).parent / "cleaned"
HISTORICAL_DIR = Path(__file__).parent / "historical"
OLDER_DIR = HISTORICAL_DIR / "older"

# Legacy default — used when summary.json has no tier/quality metadata
_LEGACY_OFFER_IDS = list(range(75, 87)) + list(range(88, 105))  # 75-86, 88-104 (no 87)


def _discover_cleaned_offer_ids() -> set[int]:
    """Scan cleaned/ for offer_*_mystery.csv and return available offer IDs."""
    ids = set()
    for f in CLEANED_DIR.glob("offer_*_mystery.csv"):
        try:
            ids.add(int(f.stem.split("_")[1]))
        except (ValueError, IndexError):
            pass
    return ids


def _effective_algorithm(algorithm: str | None) -> str:
    """The strategy to run when the caller passes no --algorithm (production default)."""
    return algorithm or DEFAULT_STRATEGY


def _offer_tier(offer_id: int) -> str:
    """Determine data tier from offer ID range."""
    if offer_id >= 64:
        return "A"
    if offer_id >= 55:
        return "B"
    if offer_id >= 49:
        return "C"
    return "D"


def _build_offer_ids(summary: dict, only_offers: str | None = None) -> list[int]:
    """
    Build the list of offer IDs to process.

    Args:
        summary: The cleaned/summary.json data
        only_offers: Optional comma-separated IDs and ranges (e.g. "55,60,70-80")

    Returns sorted list of offer IDs.
    """
    available = _discover_cleaned_offer_ids()

    if only_offers:
        ids = set()
        for part in only_offers.split(","):
            part = part.strip()
            if "-" in part:
                lo, hi = part.split("-", 1)
                ids.update(range(int(lo), int(hi) + 1))
            else:
                ids.add(int(part))
        return sorted(ids & available)

    # Default: use Tier A offers only (reliable data quality).
    # Use --only-offers to include Tier B/C/D.
    return sorted(oid for oid in available if _offer_tier(oid) == "A")


def _find_xlsx_path(offer_id: int) -> Path | None:
    """Find the XLSX file for an offer, checking both directories."""
    for d in [HISTORICAL_DIR, OLDER_DIR]:
        # Try canonical name first
        canonical = d / f"offer_{offer_id}_shopping_list.xlsx"
        if canonical.exists():
            return canonical
        # Try glob pattern for non-canonical names
        matches = list(d.glob(f"offer_{offer_id}_shopping_list*.xlsx"))
        if matches:
            return matches[0]
    return None


# Module-level OFFER_IDS — set by main() or by callers
OFFER_IDS = _LEGACY_OFFER_IDS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_summary() -> dict:
    """Load the cleaned data summary.json."""
    with open(CLEANED_DIR / "summary.json") as f:
        return json.load(f)


def load_historical_csv(offer_id: int) -> tuple[list[str], dict[int, dict[str, int]]]:
    """
    Load historical mystery CSV.

    Returns:
        (box_names, allocations) where allocations is {item_id: {box_name: qty}}
    """
    path = CLEANED_DIR / f"offer_{offer_id}_mystery.csv"
    if not path.exists():
        return [], {}

    with open(path) as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        box_names = [fn for fn in fieldnames if fn != "id"]

        allocations = {}
        for row in reader:
            item_id = int(row["id"])
            allocations[item_id] = {}
            for bn in box_names:
                qty = float(row[bn]) if row[bn] else 0
                if qty > 0:
                    allocations[item_id][bn] = qty

    return box_names, allocations


def read_xlsx_pack_overrides(offer_id: int) -> dict[int, int]:
    """
    Read Price Ea from the XLSX calculation tab for pack items.

    For lemon/lime packs: if XLSX Price Ea ≈ DB price / pack_size, the pack
    was split into singles (use per-unit price). If ≈ DB price, the whole
    pack was given (use full price).

    Returns {item_id: price_in_cents} for items needing override.
    """
    import openpyxl
    xlsx_path = _find_xlsx_path(offer_id)
    if xlsx_path is None:
        return {}

    db_items = {r["item_id"]: r for r in fetch_offer_items(offer_id)}

    # Find pack items
    pack_items = {}
    for iid, item in db_items.items():
        ps = detect_pack_size(item["item_name"])
        if ps > 1:
            pack_items[iid] = (item, ps)
    if not pack_items:
        return {}

    # Read Price Ea from penultimate sheet (or first sheet with price column)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = wb.sheetnames
    if len(sheets) < 2:
        return {}

    # Try penultimate sheet first, then others
    candidates = [sheets[-2]] + [s for s in sheets if s != sheets[-2]]
    ws = None
    id_col = price_col = None
    for sn in candidates:
        _ws = wb[sn]
        if not _ws.max_row or _ws.max_row < 2:
            continue
        headers = [cell.value for cell in list(_ws.iter_rows(min_row=1, max_row=1))[0]]
        _id_col = _price_col = None
        for i, h in enumerate(headers):
            hs = str(h).strip() if h else ""
            if hs == "ID":
                _id_col = i
            if hs in ("Price Ea", "JS Price Ea"):
                _price_col = i
        if _id_col is not None and _price_col is not None:
            ws, id_col, price_col = _ws, _id_col, _price_col
            break

    if ws is None or id_col is None or price_col is None:
        return {}

    overrides = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        iid_raw, price_raw = row[id_col], row[price_col]
        if iid_raw is None or price_raw is None:
            continue
        try:
            iid = int(iid_raw)
            xlsx_price = float(price_raw)
        except (ValueError, TypeError):
            continue
        if iid not in pack_items:
            continue

        item, ps = pack_items[iid]
        db_price = item["item_price"]  # cents
        per_unit = db_price // ps
        xlsx_cents = round(xlsx_price * 100)

        # If XLSX price ≈ full DB price, whole pack was given
        if abs(xlsx_cents - db_price) < PACK_PRICE_TOLERANCE_CENTS:
            overrides[iid] = db_price
        # If XLSX price ≈ per-unit price, pack was split (default behaviour)
        elif abs(xlsx_cents - per_unit) < PACK_PRICE_TOLERANCE_CENTS:
            pass  # no override needed, detect_pack_size handles it
        else:
            # XLSX has a custom adjusted price — use it directly
            overrides[iid] = xlsx_cents

    return overrides


def build_item_lookup(
    offer_id: int,
    price_overrides: dict[int, int] | None = None,
) -> dict[int, dict]:
    """
    Build item lookup from DB: {item_id: {name, price, category_id, fungible_group}}.

    price_overrides: optional {item_id: price_in_cents} to override DB prices
    (e.g. from XLSX Price Ea for pack items where the price differs from DB/pack_size).
    """
    categories = fetch_categories()
    db_items = fetch_offer_items(offer_id)
    lookup = {}
    for row in db_items:
        fg, degree = assign_fungible_group(row["item_name"])
        sub_cat, usage, colour, shape = assign_classification(
            row["item_name"], row["category_id"]
        )
        if price_overrides and row["item_id"] in price_overrides:
            price = price_overrides[row["item_id"]]
        else:
            pack_size = detect_pack_size(row["item_name"])
            price = row["item_price"] // pack_size
        lookup[row["item_id"]] = {
            "name": row["item_name"],
            "price": price,
            "category_id": row["category_id"],
            "category_name": categories.get(row["category_id"], "unknown").lower(),
            "size": row.get("size", 1) or 1,
            "fungible_group": fg,
            "fungible_degree": degree,
            "sub_category": sub_cat,
            "usage": usage,
            "colour": colour,
            "shape": shape,
        }
    return lookup


# ---------------------------------------------------------------------------
# Metrics computation
# ---------------------------------------------------------------------------

def compute_available_tags(
    item_lookup: dict[int, dict], preference: str | None = None,
) -> dict[str, set[str]]:
    """Compute distinct tags available across all items in the offer.

    If preference is set, filter to items matching that preference.
    """
    tags: dict[str, set[str]] = {
        "sub_category": set(), "usage": set(), "colour": set(), "shape": set(),
    }
    for info in item_lookup.values():
        if preference == "fruit_only" and info["category_id"] != CATEGORY_FRUIT:
            continue
        if preference == "veg_only" and info["category_id"] != CATEGORY_VEGETABLES:
            continue
        if info.get("sub_category"):
            tags["sub_category"].add(info["sub_category"])
        if info.get("usage"):
            tags["usage"].add(info["usage"])
        if info.get("colour"):
            tags["colour"].add(info["colour"])
        if info.get("shape"):
            tags["shape"].add(info["shape"])
    return tags


def compute_box_metrics(
    box_name: str,
    allocations: dict[int, int],  # item_id -> qty for this box
    item_lookup: dict[int, dict],
    tier: str,
    preference: str | None = None,
    available_tags: dict[str, set[str]] | None = None,
) -> dict | None:
    """
    Compute metrics for a single box.

    Returns dict with:
        tier, target_value, total_value, value_pct, unique_items,
        fruit_value, veg_value, fruit_pct, diversity_score,
        fungible_dupes, pref_violations
    Or None if no items could be resolved.
    """
    price = BOX_TIERS[tier]["price"]
    target = BOX_TIERS[tier]["target_value"]

    total_value = 0
    fruit_value = 0
    veg_value = 0
    unique_items = 0
    # Also track variety count per fungible group (for informational display)
    fungible_groups_seen: dict[str, tuple[int, float]] = {}
    pref_violations = 0
    resolved_items = 0

    # Track qty-weighted diversity tag counts for this box
    box_tag_counts: dict[str, dict[str, int]] = {
        "sub_category": {}, "usage": {}, "colour": {}, "shape": {},
    }

    # Per-item data for same-item and group-concentration penalties
    per_item_data: list[tuple[int, int, dict]] = []  # (item_id, qty, info)

    for item_id, qty in allocations.items():
        if item_id not in item_lookup:
            continue  # item no longer in DB, skip

        info = item_lookup[item_id]
        if qty <= 0:
            continue

        resolved_items += 1
        val = info["price"] * qty
        total_value += val
        unique_items += 1
        per_item_data.append((item_id, qty, info))

        if info["category_id"] == CATEGORY_FRUIT:
            fruit_value += val
        elif info["category_id"] == CATEGORY_VEGETABLES:
            veg_value += val

        # Accumulate qty-weighted diversity tag counts
        if info.get("sub_category"):
            box_tag_counts["sub_category"][info["sub_category"]] = box_tag_counts["sub_category"].get(info["sub_category"], 0) + qty
        if info.get("usage"):
            box_tag_counts["usage"][info["usage"]] = box_tag_counts["usage"].get(info["usage"], 0) + qty
        if info.get("colour"):
            box_tag_counts["colour"][info["colour"]] = box_tag_counts["colour"].get(info["colour"], 0) + qty
        if info.get("shape"):
            box_tag_counts["shape"][info["shape"]] = box_tag_counts["shape"].get(info["shape"], 0) + qty

        fg = info["fungible_group"]
        if fg:
            degree = info["fungible_degree"]
            if fg in fungible_groups_seen:
                prev_count, prev_degree = fungible_groups_seen[fg]
                fungible_groups_seen[fg] = (prev_count + 1, prev_degree)
            else:
                fungible_groups_seen[fg] = (1, degree)

        # Preference compliance
        if preference == "fruit_only" and info["category_id"] == CATEGORY_VEGETABLES:
            pref_violations += 1
        elif preference == "veg_only" and info["category_id"] == CATEGORY_FRUIT:
            pref_violations += 1

    if resolved_items == 0:
        return None

    # Informational fungible dupe counts (for display, not scoring)
    fungible_dupes = 0
    slot_dupes = 0
    bad_dupes = 0
    for fg, (count, degree) in fungible_groups_seen.items():
        dupes = max(0, count - 1)
        fungible_dupes += dupes
        if degree >= SLOT_DEGREE_THRESHOLD:
            slot_dupes += dupes
        else:
            bad_dupes += dupes

    # Same-item penalty
    same_item_penalty = 0.0
    for item_id, qty, info in per_item_data:
        item_allow = _resolve_item_allowance_from_lookup(info, tier)
        excess = max(0, qty - item_allow)
        if excess > 0:
            same_item_penalty += excess * info["price"] * SAME_ITEM_MULTIPLIER / 100.0

    # Group concentration penalty (uses min(qty, item_allowance) for group load)
    group_loads: dict[str, tuple[float, float]] = {}  # group -> (load, degree)
    for item_id, qty, info in per_item_data:
        fg = info["fungible_group"]
        if not fg:
            continue
        degree = info["fungible_degree"]
        item_allow = _resolve_item_allowance_from_lookup(info, tier)
        capped = min(qty, item_allow)
        if fg in group_loads:
            prev_load, prev_deg = group_loads[fg]
            group_loads[fg] = (prev_load + capped, prev_deg)
        else:
            group_loads[fg] = (capped, degree)

    group_concentration_penalty = 0.0
    for fg, (load, degree) in group_loads.items():
        if fg in GROUP_ALLOWANCES:
            ga = GROUP_ALLOWANCES[fg].get(tier, load)
            excess = max(0, load - ga)
            if excess > 0:
                group_concentration_penalty += (excess ** GROUP_QTY_EXPONENT) * degree * GROUP_CONCENTRATION_MULTIPLIER

    # Max value share penalty
    max_value_share = 0.0
    if total_value > 0:
        for item_id, qty, info in per_item_data:
            share = info["price"] * qty / total_value
            if share > max_value_share:
                max_value_share = share
    mvs_excess = max(0.0, max_value_share - MAX_VALUE_SHARE_THRESHOLD)
    max_value_share_penalty = mvs_excess * MAX_VALUE_SHARE_MULTIPLIER

    # Size floor penalty
    total_size_points = 0
    for item_id, qty, info in per_item_data:
        total_size_points += info.get("size", 1) * qty
    sf_target = SIZE_FLOOR_TARGETS.get(tier, 0)
    size_floor_penalty = max(0, sf_target - total_size_points) * SIZE_FLOOR_MULTIPLIER

    total_fv = fruit_value + veg_value
    fruit_pct = (fruit_value / total_fv * 100) if total_fv > 0 else 0.0

    # Compute diversity score using effective number of species (1/HHI)
    diversity_score = 0.0
    if available_tags:
        for dim, weight in DIVERSITY_WEIGHTS.items():
            n_available = len(available_tags.get(dim, set()))
            dim_counts = box_tag_counts.get(dim, {})
            if n_available > 0 and dim_counts:
                total_qty = sum(dim_counts.values())
                hhi = sum((q / total_qty) ** 2 for q in dim_counts.values())
                eff_species = 1.0 / hhi
                diversity_score += weight * min(eff_species / n_available, 1.0)
            elif n_available == 0:
                diversity_score += weight
            # else: no items with this dimension tag → 0 contribution
    else:
        diversity_score = DIVERSITY_FALLBACK_SCORE  # no reference = neutral

    return {
        "box_name": box_name,
        "tier": tier,
        "target_value": target,
        "total_value": total_value,
        "value_pct": (total_value / price * 100) if price > 0 else 0.0,
        "unique_items": unique_items,
        "fruit_value": fruit_value,
        "veg_value": veg_value,
        "fruit_pct": fruit_pct,
        "diversity_score": diversity_score,
        "same_item_penalty": same_item_penalty,
        "group_concentration_penalty": group_concentration_penalty,
        "max_value_share": max_value_share,
        "max_value_share_penalty": max_value_share_penalty,
        "total_size_points": total_size_points,
        "size_floor_penalty": size_floor_penalty,
        "fungible_dupes": fungible_dupes,
        "slot_dupes": slot_dupes,
        "bad_dupes": bad_dupes,
        "pref_violations": pref_violations,
    }


def compute_composite_score(metrics: list[dict], params: dict | None = None) -> dict:
    """
    Compute composite quality score from box metrics.

    Returns dict with total score and individual penalty breakdowns.
    Score = 100 - penalties. Higher is better.
    """
    if not metrics:
        return {"score": 0.0, "value_pen": 0.0, "si_pen": 0.0, "gc_pen": 0.0,
                "diversity_pen": 0.0, "mvs_pen": 0.0, "sf_pen": 0.0,
                "pref_pen": 0.0}

    n = len(metrics)

    div_mult = (params or {}).get("diversity_penalty_multiplier", DIVERSITY_PENALTY_MULTIPLIER)

    # 1. Value penalty (per box, averaged)
    value_penalties = [value_penalty(m["value_pct"], params) for m in metrics]
    avg_value_pen = sum(value_penalties) / n

    # 2. Same-item penalty (per box, averaged — already includes multiplier)
    avg_si_pen = sum(m.get("same_item_penalty", 0.0) for m in metrics) / n

    # 3. Group concentration penalty (per box, averaged — already includes multiplier)
    avg_gc_pen = sum(m.get("group_concentration_penalty", 0.0) for m in metrics) / n

    # 4. Diversity penalty (per box, averaged)
    avg_diversity_pen = sum((1.0 - m["diversity_score"]) * div_mult for m in metrics) / n

    # 5. Max value share penalty (per box, averaged — already includes multiplier)
    avg_mvs_pen = sum(m.get("max_value_share_penalty", 0.0) for m in metrics) / n

    # 6. Size floor penalty (per box, averaged — already includes multiplier)
    avg_sf_pen = sum(m.get("size_floor_penalty", 0.0) for m in metrics) / n

    # 7. Preference violations (hard penalty)
    pref_pen = sum(m["pref_violations"] for m in metrics) * PREF_VIOLATION_PENALTY

    score = MAX_COMPOSITE_SCORE - avg_value_pen - avg_si_pen - avg_gc_pen - avg_diversity_pen - avg_mvs_pen - avg_sf_pen - pref_pen
    return {
        "score": score,
        "value_pen": avg_value_pen,
        "si_pen": avg_si_pen,
        "gc_pen": avg_gc_pen,
        "diversity_pen": avg_diversity_pen,
        "mvs_pen": avg_mvs_pen,
        "sf_pen": avg_sf_pen,
        "pref_pen": pref_pen,
    }


# ---------------------------------------------------------------------------
# Process offers
# ---------------------------------------------------------------------------

def _process_offer_quiet(
    offer_id: int,
    summary: dict,
    algorithm: str | None = None,
    bootstrap_allocations: list[dict[int, int]] | None = None,
) -> tuple[list[dict], list[dict]] | None:
    """
    Process a single offer without printing. Pure computation.

    Returns (manual_metrics, algo_metrics) or None if the offer can't be processed.
    """
    item_lookup = build_item_lookup(offer_id)
    if not item_lookup:
        return None

    # Build a separate lookup for manual scoring with XLSX pack-price overrides
    pack_overrides = read_xlsx_pack_overrides(offer_id)
    manual_item_lookup = (
        build_item_lookup(offer_id, price_overrides=pack_overrides)
        if pack_overrides else item_lookup
    )

    box_names, hist_allocs = load_historical_csv(offer_id)
    if not box_names:
        return None

    from allocator.config import SKIP_COLUMN_IDENTIFIERS, STAFF_IDENTIFIERS
    box_names = [bn for bn in box_names
                 if bn not in DONATION_IDENTIFIERS
                 and bn not in SKIP_COLUMN_IDENTIFIERS
                 and bn not in STAFF_IDENTIFIERS]

    buyers_db = fetch_mystery_box_buyers(offer_id)
    buyer_prefs = {}
    for buyer in buyers_db:
        email = buyer["user_email"]
        opt = buyer.get("selected_option") or ""
        if "no veg" in opt.lower():
            buyer_prefs[email] = "fruit_only"
        elif "no fruit" in opt.lower():
            buyer_prefs[email] = "veg_only"

    # Precompute available tags for each preference variant
    avail_tags_all = compute_available_tags(item_lookup)
    avail_tags_fruit = compute_available_tags(item_lookup, preference="fruit_only")
    avail_tags_veg = compute_available_tags(item_lookup, preference="veg_only")

    def _tags_for_pref(pref):
        if pref == "fruit_only":
            return avail_tags_fruit
        if pref == "veg_only":
            return avail_tags_veg
        return avail_tags_all

    manual_metrics = []
    for bn in box_names:
        tier = infer_box_tier(offer_id, bn, summary)
        if tier is None:
            continue
        box_allocs = {}
        for item_id, per_box in hist_allocs.items():
            qty = per_box.get(bn, 0)
            if qty > 0:
                box_allocs[item_id] = qty

        pref = buyer_prefs.get(bn)
        m = compute_box_metrics(bn, box_allocs, manual_item_lookup, tier, preference=pref,
                                available_tags=_tags_for_pref(pref))
        if m:
            m["source"] = "manual"
            manual_metrics.append(m)

    # "manual" pseudo-strategy: score the historical CSV directly
    if algorithm == "manual":
        algo_metrics = []
        for m in manual_metrics:
            am = dict(m)
            am["source"] = "algorithm"
            algo_metrics.append(am)
        return manual_metrics, algo_metrics, []

    xlsx_path = _find_xlsx_path(offer_id)
    if xlsx_path is None:
        # No XLSX → can only score historical CSV, skip algorithm comparison
        return manual_metrics, [], []

    try:
        kwargs = {}
        if algorithm:
            kwargs["strategy"] = algorithm
        if bootstrap_allocations is not None:
            kwargs["bootstrap_allocations"] = bootstrap_allocations
        result = allocate(offer_id, xlsx_path, **kwargs)
    except Exception as e:
        print(f"  [ERROR] Algorithm failed for offer {offer_id}: {e}", file=sys.stderr)
        return manual_metrics, [], []

    algo_metrics = []
    box_allocations = []
    for box in result.boxes:
        box_allocations.append(dict(box.allocations))
        m = compute_box_metrics(
            box.name,
            box.allocations,
            item_lookup,
            box.tier,
            preference=box.preference,
            available_tags=_tags_for_pref(box.preference),
        )
        if m:
            m["source"] = "algorithm"
            algo_metrics.append(m)

    return manual_metrics, algo_metrics, box_allocations


def process_offer(offer_id: int, summary: dict, algorithm: str | None = None, verbose: bool = True) -> tuple[list[dict], list[dict]] | None:
    """
    Process a single offer: compute metrics for both manual and algorithm.

    Returns (manual_metrics, algo_metrics) or None if the offer can't be processed.
    """
    print(f"\n{'='*70}")
    print(f"  OFFER {offer_id}")
    print(f"{'='*70}")

    result = _process_offer_quiet(offer_id, summary, algorithm=algorithm)
    if result is None:
        print(f"  [SKIP] Offer {offer_id}")
        return None

    manual_metrics, algo_metrics, _box_allocs = result

    print(f"  Manual boxes: {len(manual_metrics)},  Algorithm boxes: {len(algo_metrics)}")

    if verbose:
        box_header = (f"  {'Box':<35} {'Tier':<7} {'Value':>8} {'Target':>8} {'%Tgt':>6} "
                      f"{'Items':>5} {'Fr%':>5} {'Diver':>5} {'FDup':>4} {'BDup':>4} {'PVio':>4}")
        print(f"\n  --- Manual (historical) per-box ---")
        print(box_header)
        for m in manual_metrics:
            print(f"  {m['box_name'][:34]:<35} {m['tier']:<7} "
                  f"${m['total_value']/100:>7.2f} ${m['target_value']/100:>7.2f} "
                  f"{m['value_pct']:>5.1f}% {m['unique_items']:>5} "
                  f"{m['fruit_pct']:>4.1f}% {m['diversity_score']:>5.2f} "
                  f"{m['fungible_dupes']:>4} {m['bad_dupes']:>4} {m['pref_violations']:>4}")

        print(f"\n  --- Algorithm per-box ---")
        print(box_header)
        for m in algo_metrics:
            print(f"  {m['box_name'][:34]:<35} {m['tier']:<7} "
                  f"${m['total_value']/100:>7.2f} ${m['target_value']/100:>7.2f} "
                  f"{m['value_pct']:>5.1f}% {m['unique_items']:>5} "
                  f"{m['fruit_pct']:>4.1f}% {m['diversity_score']:>5.2f} "
                  f"{m['fungible_dupes']:>4} {m['bad_dupes']:>4} {m['pref_violations']:>4}")

    return manual_metrics, algo_metrics


def compute_averages(metrics: list[dict]) -> dict:
    """Compute average metrics across all boxes."""
    if not metrics:
        return {}

    n = len(metrics)
    avg_value_pct = sum(m["value_pct"] for m in metrics) / n
    return {
        "count": n,
        "avg_value": sum(m["total_value"] for m in metrics) / n,
        "avg_value_pct": avg_value_pct,
        "avg_unique_items": sum(m["unique_items"] for m in metrics) / n,
        "avg_fruit_pct": sum(m["fruit_pct"] for m in metrics) / n,
        "avg_diversity_score": sum(m["diversity_score"] for m in metrics) / n,
        "avg_fungible_dupes": sum(m["fungible_dupes"] for m in metrics) / n,
        "avg_slot_dupes": sum(m["slot_dupes"] for m in metrics) / n,
        "avg_bad_dupes": sum(m["bad_dupes"] for m in metrics) / n,
        "avg_pref_violations": sum(m["pref_violations"] for m in metrics) / n,
        "std_value_pct": (
            sum((m["value_pct"] - avg_value_pct) ** 2 for m in metrics) / n
        ) ** 0.5,
        "pct_above_target": sum(1 for m in metrics if m["total_value"] >= m["target_value"]) / n * 100,
        "pct_within_15": sum(
            1 for m in metrics
            if 100 <= m["value_pct"] <= 130
        ) / n * 100,
        "composite": compute_composite_score(metrics),
    }


def compute_tier_averages(metrics: list[dict]) -> dict[str, dict]:
    """Compute average metrics grouped by tier."""
    by_tier = defaultdict(list)
    for m in metrics:
        by_tier[m["tier"]].append(m)
    return {tier: compute_averages(ms) for tier, ms in sorted(by_tier.items())}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _process_offer_worker(args):
    """Worker for parallel execution. Must be top-level for pickling."""
    offer_id, algorithm, summary = args[:3]
    bootstrap_allocations = args[3] if len(args) > 3 else None
    try:
        result = _process_offer_quiet(
            offer_id, summary, algorithm=algorithm,
            bootstrap_allocations=bootstrap_allocations,
        )
        if result is None:
            return (offer_id, algorithm, None, None, None)
        return (offer_id, algorithm, result[0], result[1], result[2])
    except Exception as e:
        print(f"[ERROR] offer={offer_id} strategy={algorithm}: {e}", file=sys.stderr)
        return (offer_id, algorithm, None, None, None)


def run_all_strategies_parallel(summary, max_workers=None):
    """Run all strategies in parallel across all offers.

    local-search bootstraps from discard-worst, so we run in two phases to
    avoid computing discard-worst twice per offer:
      Phase 1: all strategies except local-search (captures discard-worst allocations)
      Phase 2: local-search with pre-computed discard-worst allocations
    """
    from concurrent.futures import ProcessPoolExecutor, as_completed

    strategies = list_strategies(include_baselines=True) + ["manual"]
    phase1_strategies = [s for s in strategies if s != "local-search"]
    n_offers = len(OFFER_IDS)
    n_tasks = len(strategies) * n_offers
    t0 = time.monotonic()

    strategy_algo = {s: [] for s in strategies}
    manual_seen = set()
    all_manual = []
    done_per_strat = {s: 0 for s in strategies}

    # Capture discard-worst box allocations per offer for local-search bootstrap
    dw_allocations: dict[int, list[dict[int, int]]] = {}

    print(f"  Running {n_tasks} tasks ({len(strategies)} strategies × {n_offers} offers) "
          f"with {max_workers or os.cpu_count()} workers...")

    def _collect(future):
        offer_id, strat, manual, algo, box_allocs = future.result()
        if algo:
            strategy_algo[strat].extend(algo)
        if manual and offer_id not in manual_seen:
            manual_seen.add(offer_id)
            all_manual.extend(manual)
        if strat == "discard-worst" and box_allocs:
            dw_allocations[offer_id] = box_allocs
        done_per_strat[strat] += 1
        if done_per_strat[strat] == n_offers:
            elapsed = time.monotonic() - t0
            score = compute_composite_score(strategy_algo[strat])["score"]
            print(f"  {strat:<20} score={score:.1f}  ({elapsed:.0f}s)")

    # Phase 1: everything except local-search
    phase1_tasks = [(oid, strat, summary) for strat in phase1_strategies for oid in OFFER_IDS]
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_process_offer_worker, t): t for t in phase1_tasks}
        for future in as_completed(futures):
            _collect(future)

    # Phase 2: local-search with pre-computed discard-worst allocations
    if "local-search" in strategies:
        phase2_tasks = [
            (oid, "local-search", summary, dw_allocations.get(oid))
            for oid in OFFER_IDS
        ]
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(_process_offer_worker, t): t for t in phase2_tasks}
            for future in as_completed(futures):
                _collect(future)

    return strategy_algo, all_manual


def run_single_strategy_parallel(algorithm, summary, max_workers=None):
    """Run a single strategy in parallel across all offers.

    Returns (all_manual, all_algo, per_offer) where per_offer is
    {offer_id: (manual_metrics, algo_metrics)}.
    """
    from concurrent.futures import ProcessPoolExecutor

    alg = _effective_algorithm(algorithm)
    tasks = [(oid, alg, summary) for oid in OFFER_IDS]

    all_manual, all_algo = [], []
    per_offer = {}

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        for offer_id, _, manual, algo, _ in executor.map(
            _process_offer_worker, tasks, chunksize=1
        ):
            if manual:
                all_manual.extend(manual)
            if algo:
                all_algo.extend(algo)
            if manual or algo:
                per_offer[offer_id] = (manual or [], algo or [])

    return all_manual, all_algo, per_offer


def run_single_strategy(algorithm: str | None, summary: dict, verbose: bool = True) -> tuple[list[dict], list[dict], dict]:
    """Run comparison for a single strategy across all offers.

    Returns (all_manual, all_algo, per_offer).
    """
    all_manual = []
    all_algo = []
    per_offer = {}

    for offer_id in OFFER_IDS:
        result = process_offer(offer_id, summary, algorithm=algorithm, verbose=verbose)
        if result is None:
            continue
        manual_metrics, algo_metrics = result
        all_manual.extend(manual_metrics)
        all_algo.extend(algo_metrics)
        per_offer[offer_id] = (manual_metrics, algo_metrics)

    return all_manual, all_algo, per_offer


def print_leaderboard(results: dict[str, dict]):
    """Print strategy leaderboard sorted by composite score."""
    from rich.console import Console
    from rich.table import Table
    from rich import box as rich_box

    console = Console()
    ranked = sorted(results.items(), key=lambda x: x[1]["composite"]["score"], reverse=True)

    table = Table(
        title=f"Canonical vs baselines  (across {len(OFFER_IDS)} offers)",
        box=rich_box.SIMPLE_HEAD,
        show_footer=False,
        title_style="bold",
    )
    table.add_column("Rank", style="bold cyan", justify="right", width=4)
    table.add_column("Strategy", min_width=18)
    table.add_column("Score", justify="right", style="bold green")
    table.add_column("Value", justify="right")
    table.add_column("SameI", justify="right")
    table.add_column("GrpC", justify="right")
    table.add_column("Diver", justify="right")
    table.add_column("MVS", justify="right")
    table.add_column("SzFlr", justify="right")
    table.add_column("Pref", justify="right")

    for i, (name, avg) in enumerate(ranked, 1):
        c = avg["composite"]
        score_str = f"{c['score']:.1f}"
        rank_str = f"{i}."
        table.add_row(
            rank_str,
            name,
            score_str,
            f"{-c['value_pen']:+.1f}",
            f"{-c.get('si_pen', 0):+.1f}",
            f"{-c.get('gc_pen', 0):+.1f}",
            f"{-c['diversity_pen']:+.1f}",
            f"{-c.get('mvs_pen', 0):+.1f}",
            f"{-c.get('sf_pen', 0):+.1f}",
            f"{-c['pref_pen']:+.1f}",
        )

    console.print()
    console.print(table)
    console.print("  Score = 100 minus penalties.  [green]>80[/green] = good, "
                  "[yellow]>60[/yellow] = acceptable, [red]<60[/red] = poor.")
    console.print("  Penalty columns show points subtracted from 100 per dimension.")
    console.print()


def print_detail(algorithm: str, per_offer: dict, all_algo: list[dict]):
    """Print per-offer breakdown and value distribution, write JSON dump."""
    alg_name = _effective_algorithm(algorithm)

    # --- Per-offer table (worst first) ---
    offer_rows = []
    for offer_id, (manual, algo) in per_offer.items():
        if not algo:
            continue
        comp = compute_composite_score(algo)
        offer_rows.append((offer_id, len(algo), comp))

    offer_rows.sort(key=lambda r: r[2]["score"])

    print(f"\n{'='*80}")
    print(f"  PER-OFFER BREAKDOWN — {alg_name} (worst first)")
    print(f"{'='*80}")
    print(f"  {'Offer':>5} {'Boxes':>5} {'Score':>7} {'Value':>7} {'SameI':>7} "
          f"{'GrpC':>7} {'Diver':>7} {'MVS':>7} {'SzFlr':>7} {'Pref':>7}")
    print(f"  {'-'*70}")

    for offer_id, n_boxes, comp in offer_rows:
        print(f"  {offer_id:>5} {n_boxes:>5} {comp['score']:>7.1f} "
              f"{-comp['value_pen']:>+7.1f} {-comp.get('si_pen', 0):>+7.1f} "
              f"{-comp.get('gc_pen', 0):>+7.1f} {-comp['diversity_pen']:>+7.1f} "
              f"{-comp.get('mvs_pen', 0):>+7.1f} {-comp.get('sf_pen', 0):>+7.1f} "
              f"{-comp['pref_pen']:>+7.1f}")

    # --- Value distribution ---
    buckets = [
        ("<100%", lambda v: v < 100),
        ("100-110%", lambda v: 100 <= v < 110),
        ("110-114%", lambda v: 110 <= v < 114),
        ("114-117%", lambda v: 114 <= v < 117),  # sweet spot
        ("117-120%", lambda v: 117 <= v < 120),
        ("120-130%", lambda v: 120 <= v < 130),
        (">=130%", lambda v: v >= 130),
    ]

    n_total = len(all_algo)
    print(f"\n{'='*80}")
    print(f"  VALUE % DISTRIBUTION — {alg_name} ({n_total} boxes)")
    print(f"{'='*80}")

    for label, test in buckets:
        count = sum(1 for m in all_algo if test(m["value_pct"]))
        pct = count / n_total * 100 if n_total else 0
        bar = "#" * int(pct / 2)
        marker = " <-- sweet spot" if "114-117" in label else ""
        print(f"  {label:>10}: {count:>4} ({pct:>5.1f}%) {bar}{marker}")

    # --- Worst boxes across all offers ---
    print(f"\n{'='*80}")
    print(f"  WORST 20 BOXES BY VALUE PENALTY — {alg_name}")
    print(f"{'='*80}")
    print(f"  {'Offer':>5} {'Box':<30} {'Tier':<7} {'Value%':>7} {'Penalty':>8} "
          f"{'BDup':>5} {'Diver':>6}")
    print(f"  {'-'*73}")

    # Annotate each box metric with its offer_id
    annotated = []
    for offer_id, (manual, algo) in per_offer.items():
        for m in algo:
            pen = value_penalty(m["value_pct"])
            annotated.append((offer_id, m, pen))

    annotated.sort(key=lambda x: -x[2])
    for offer_id, m, pen in annotated[:20]:
        print(f"  {offer_id:>5} {m['box_name'][:29]:<30} {m['tier']:<7} "
              f"{m['value_pct']:>6.1f}% {pen:>+8.1f} "
              f"{m['bad_dupes']:>5} {m['diversity_score']:>6.2f}")

    # --- JSON dump ---
    dump = {
        "strategy": alg_name,
        "aggregate": compute_composite_score(all_algo),
        "offers": {},
    }
    for offer_id, (manual, algo) in sorted(per_offer.items()):
        dump["offers"][offer_id] = {
            "composite": compute_composite_score(algo),
            "manual_composite": compute_composite_score(manual),
            "boxes": algo,
            "manual_boxes": manual,
        }

    dump_path = Path(__file__).parent / "output" / f"detail_{alg_name}.json"
    dump_path.parent.mkdir(exist_ok=True)
    with open(dump_path, "w") as f:
        json.dump(dump, f, indent=2, default=str)
    print(f"\n  Full detail written to {dump_path}")
    print()


def _box_main_issue(value_pen, gq_pen, div_pen):
    """Classify the main issue(s) for a box based on penalty magnitudes."""
    issues = []
    if value_pen >= 2.0:
        issues.append("value")
    if gq_pen >= 1.0:
        issues.append("group_qty")
    if div_pen >= 5.0:
        issues.append("diversity")
    return ", ".join(issues) if issues else "ok"


def write_csv(label: str, per_offer: dict, source: str):
    """Write per-box metrics CSV to output/.

    source: "manual" or "algo" — selects which side of per_offer to use.
    """
    out_dir = Path(__file__).parent / "output"
    out_dir.mkdir(exist_ok=True)
    csv_path = out_dir / f"{label}_scores.csv"

    fieldnames = [
        "offer", "box", "tier", "target_value_cents", "total_value_cents",
        "value_pct", "score", "value_penalty", "same_item_penalty",
        "group_concentration_penalty", "diversity_penalty",
        "max_value_share_penalty", "size_floor_penalty",
        "diversity_score", "max_value_share",
        "bad_dupes", "fungible_dupes", "slot_dupes", "unique_items",
        "fruit_pct", "pref_violations", "main_issue",
    ]

    rows = []
    for offer_id in sorted(per_offer.keys(), key=int):
        manual, algo = per_offer[offer_id]
        boxes = manual if source == "manual" else algo
        for m in boxes:
            vp = value_penalty(m["value_pct"])
            si_pen = m.get("same_item_penalty", 0.0)
            gc_pen = m.get("group_concentration_penalty", 0.0)
            dv = (1.0 - m["diversity_score"]) * DIVERSITY_PENALTY_MULTIPLIER
            mvs_pen = m.get("max_value_share_penalty", 0.0)
            sf_pen = m.get("size_floor_penalty", 0.0)
            score = MAX_COMPOSITE_SCORE - vp - si_pen - gc_pen - dv - mvs_pen - sf_pen
            rows.append({
                "offer": offer_id,
                "box": m["box_name"],
                "tier": m["tier"],
                "target_value_cents": m["target_value"],
                "total_value_cents": m["total_value"],
                "value_pct": round(m["value_pct"], 2),
                "score": round(score, 2),
                "value_penalty": round(vp, 2),
                "same_item_penalty": round(si_pen, 2),
                "group_concentration_penalty": round(gc_pen, 1),
                "diversity_penalty": round(dv, 2),
                "max_value_share_penalty": round(mvs_pen, 2),
                "size_floor_penalty": round(sf_pen, 2),
                "diversity_score": round(m["diversity_score"], 4),
                "max_value_share": round(m.get("max_value_share", 0.0), 4),
                "bad_dupes": m["bad_dupes"],
                "fungible_dupes": m["fungible_dupes"],
                "slot_dupes": m["slot_dupes"],
                "unique_items": m["unique_items"],
                "fruit_pct": round(m["fruit_pct"], 1),
                "pref_violations": m["pref_violations"],
                "main_issue": _box_main_issue(vp, gc_pen, dv),
            })

    with open(csv_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  CSV written to {csv_path} ({len(rows)} boxes)")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Compare algorithm vs historical allocations")
    parser.add_argument("--algorithm", default=None, help="Allocation algorithm (default: ilp-optimal)")
    parser.add_argument("--all-strategies", action="store_true",
                        help="Run all registered strategies and print a ranked benchmark")
    parser.add_argument("--detail", action="store_true",
                        help="Print per-offer breakdown and write detailed JSON")
    parser.add_argument("--csv", action="store_true",
                        help="Write per-box CSV to output/ (implies --detail)")
    parser.add_argument("--workers", type=int, default=None,
                        help="Parallel workers (default: cpu_count)")
    parser.add_argument("--sequential", action="store_true",
                        help="Disable parallelism")
    parser.add_argument("--only-offers", type=str, default=None,
                        help="Comma-separated offer IDs/ranges (e.g. '55,60,70-80')")
    args = parser.parse_args()
    if args.csv:
        args.detail = True

    # --verbose not a flag yet, but process_offer uses verbose=True by default
    # for single-strategy mode. Parallel mode suppresses per-offer output.
    if args.sequential or (args.workers is not None and args.workers <= 1):
        max_workers = 1
    elif args.workers is not None:
        max_workers = max(1, args.workers)
    else:
        max_workers = None  # cpu_count default

    is_parallel = max_workers is None or max_workers > 1

    summary = load_summary()

    # Set global OFFER_IDS from --only-offers or summary.json
    global OFFER_IDS
    OFFER_IDS = _build_offer_ids(summary, only_offers=args.only_offers)

    if args.all_strategies:
        t0 = time.monotonic()

        if is_parallel:
            strategy_algo, all_manual = run_all_strategies_parallel(summary, max_workers=max_workers)

            strategy_results = {}
            for strat_name, metrics in strategy_algo.items():
                if not metrics:
                    continue
                strategy_results[strat_name] = compute_averages(metrics)

        else:
            strategies = list_strategies(include_baselines=True) + ["manual"]
            strategy_results = {}

            for strat_name in strategies:
                print(f"\n{'#'*70}")
                print(f"  STRATEGY: {strat_name}")
                print(f"{'#'*70}")

                all_manual, all_algo, _ = run_single_strategy(strat_name, summary, verbose=False)

                if not all_algo:
                    print(f"  [SKIP] No results for {strat_name}")
                    continue

                alg_avg = compute_averages(all_algo)
                strategy_results[strat_name] = alg_avg

                print(f"  Boxes: {alg_avg['count']}, "
                      f"Avg value%: {alg_avg['avg_value_pct']:.1f}%, "
                      f"Composite: {alg_avg['composite']['score']:.1f}")

        elapsed = time.monotonic() - t0
        print_leaderboard(strategy_results)
        print(f"  Completed in {elapsed:.1f}s")
        return

    algorithm = args.algorithm

    if is_parallel:
        all_manual, all_algo, per_offer = run_single_strategy_parallel(algorithm, summary, max_workers=max_workers)
    else:
        all_manual, all_algo, per_offer = run_single_strategy(algorithm, summary)

    if not all_manual and not all_algo:
        print("\nNo offers could be processed.")
        sys.exit(1)

    # --- Overall summary ---
    print(f"\n{'='*70}")
    n_processed = len(per_offer) if per_offer else len(OFFER_IDS)
    print(f"  AGGREGATE COMPARISON ({n_processed} offers)")
    print(f"{'='*70}")

    man_avg = compute_averages(all_manual)
    alg_avg = compute_averages(all_algo)

    has_algo = bool(alg_avg)

    # When no algorithm results, use zeros for display
    if not has_algo:
        _zero_comp = {"score": 0.0, "value_pen": 0.0, "si_pen": 0.0, "gc_pen": 0.0,
                      "diversity_pen": 0.0, "mvs_pen": 0.0, "sf_pen": 0.0,
                      "pref_pen": 0.0}
        alg_avg = {
            "count": 0, "avg_value": 0, "avg_value_pct": 0, "avg_unique_items": 0,
            "avg_fruit_pct": 0, "avg_diversity_score": 0, "avg_fungible_dupes": 0,
            "avg_slot_dupes": 0, "avg_bad_dupes": 0, "avg_pref_violations": 0,
            "std_value_pct": 0, "pct_above_target": 0, "pct_within_15": 0,
            "composite": _zero_comp,
        }

    header = (
        f"  {'Metric':<30} {'Manual':>12} {'Algorithm':>12} {'Diff':>10}"
    )
    sep = "  " + "-" * 66

    print(f"\n  Overall averages across ALL boxes:")
    print(header)
    print(sep)

    if not man_avg:
        print("  No manual metrics available.")
        return

    man_comp = man_avg["composite"]
    alg_comp = alg_avg["composite"]

    if has_algo:
        rows = [
            ("Total boxes", f"{man_avg['count']}", f"{alg_avg['count']}", ""),
            (
                "COMPOSITE SCORE",
                f"{man_comp['score']:.1f}",
                f"{alg_comp['score']:.1f}",
                f"{alg_comp['score']-man_comp['score']:+.1f}",
            ),
            (
                "Avg value ($)",
                f"${man_avg['avg_value']/100:.2f}",
                f"${alg_avg['avg_value']/100:.2f}",
                f"${(alg_avg['avg_value']-man_avg['avg_value'])/100:+.2f}",
            ),
            (
                "Avg value % of price",
                f"{man_avg['avg_value_pct']:.1f}%",
                f"{alg_avg['avg_value_pct']:.1f}%",
                f"{alg_avg['avg_value_pct']-man_avg['avg_value_pct']:+.1f}pp",
            ),
            (
                "Std dev of value %",
                f"{man_avg['std_value_pct']:.1f}",
                f"{alg_avg['std_value_pct']:.1f}",
                f"{alg_avg['std_value_pct']-man_avg['std_value_pct']:+.1f}",
            ),
            (
                "% boxes >= target",
                f"{man_avg['pct_above_target']:.0f}%",
                f"{alg_avg['pct_above_target']:.0f}%",
                f"{alg_avg['pct_above_target']-man_avg['pct_above_target']:+.0f}pp",
            ),
            (
                "% boxes in 100-130% range",
                f"{man_avg['pct_within_15']:.0f}%",
                f"{alg_avg['pct_within_15']:.0f}%",
                f"{alg_avg['pct_within_15']-man_avg['pct_within_15']:+.0f}pp",
            ),
            (
                "Avg unique items / box",
                f"{man_avg['avg_unique_items']:.1f}",
                f"{alg_avg['avg_unique_items']:.1f}",
                f"{alg_avg['avg_unique_items']-man_avg['avg_unique_items']:+.1f}",
            ),
            (
                "Avg fruit %",
                f"{man_avg['avg_fruit_pct']:.1f}%",
                f"{alg_avg['avg_fruit_pct']:.1f}%",
                f"{alg_avg['avg_fruit_pct']-man_avg['avg_fruit_pct']:+.1f}pp",
            ),
            (
                "Avg diversity score",
                f"{man_avg['avg_diversity_score']:.3f}",
                f"{alg_avg['avg_diversity_score']:.3f}",
                f"{alg_avg['avg_diversity_score']-man_avg['avg_diversity_score']:+.3f}",
            ),
            (
                "Avg fungible dupes / box",
                f"{man_avg['avg_fungible_dupes']:.2f}",
                f"{alg_avg['avg_fungible_dupes']:.2f}",
                f"{alg_avg['avg_fungible_dupes']-man_avg['avg_fungible_dupes']:+.2f}",
            ),
            (
                "  (slot dupes — ok)",
                f"{man_avg['avg_slot_dupes']:.2f}",
                f"{alg_avg['avg_slot_dupes']:.2f}",
                f"{alg_avg['avg_slot_dupes']-man_avg['avg_slot_dupes']:+.2f}",
            ),
            (
                "  (bad dupes — penalised)",
                f"{man_avg['avg_bad_dupes']:.2f}",
                f"{alg_avg['avg_bad_dupes']:.2f}",
                f"{alg_avg['avg_bad_dupes']-man_avg['avg_bad_dupes']:+.2f}",
            ),
            (
                "Avg pref violations / box",
                f"{man_avg['avg_pref_violations']:.2f}",
                f"{alg_avg['avg_pref_violations']:.2f}",
                f"{alg_avg['avg_pref_violations']-man_avg['avg_pref_violations']:+.2f}",
            ),
        ]
    else:
        na = "n/a"
        rows = [
            ("Total boxes", f"{man_avg['count']}", na, ""),
            ("COMPOSITE SCORE", f"{man_comp['score']:.1f}", na, ""),
            ("Avg value ($)", f"${man_avg['avg_value']/100:.2f}", na, ""),
            ("Avg value % of price", f"{man_avg['avg_value_pct']:.1f}%", na, ""),
            ("Std dev of value %", f"{man_avg['std_value_pct']:.1f}", na, ""),
            ("% boxes >= target", f"{man_avg['pct_above_target']:.0f}%", na, ""),
            ("% boxes in 100-130% range", f"{man_avg['pct_within_15']:.0f}%", na, ""),
            ("Avg unique items / box", f"{man_avg['avg_unique_items']:.1f}", na, ""),
            ("Avg fruit %", f"{man_avg['avg_fruit_pct']:.1f}%", na, ""),
            ("Avg diversity score", f"{man_avg['avg_diversity_score']:.3f}", na, ""),
            ("Avg fungible dupes / box", f"{man_avg['avg_fungible_dupes']:.2f}", na, ""),
            ("  (slot dupes — ok)", f"{man_avg['avg_slot_dupes']:.2f}", na, ""),
            ("  (bad dupes — penalised)", f"{man_avg['avg_bad_dupes']:.2f}", na, ""),
            ("Avg pref violations / box", f"{man_avg['avg_pref_violations']:.2f}", na, ""),
        ]

    for label, man_val, alg_val, diff in rows:
        print(f"  {label:<30} {man_val:>12} {alg_val:>12} {diff:>10}")

    # --- Per-tier breakdown ---
    man_by_tier = compute_tier_averages(all_manual)
    alg_by_tier = compute_tier_averages(all_algo)

    all_tiers = sorted(set(list(man_by_tier.keys()) + list(alg_by_tier.keys())),
                       key=lambda t: {"small": 0, "medium": 1, "large": 2}.get(t, 3))

    for tier in all_tiers:
        mt = man_by_tier.get(tier)
        at = alg_by_tier.get(tier)
        if not mt and not at:
            continue

        target_val = BOX_TIERS[tier]["target_value"]
        print(f"\n  --- {tier.upper()} tier (target: ${target_val/100:.2f}) ---")
        print(header)
        print(sep)

        if mt and at:
            tier_rows = [
                ("Boxes", f"{mt['count']}", f"{at['count']}", ""),
                (
                    "Composite score",
                    f"{mt['composite']['score']:.1f}",
                    f"{at['composite']['score']:.1f}",
                    f"{at['composite']['score']-mt['composite']['score']:+.1f}",
                ),
                (
                    "Avg value ($)",
                    f"${mt['avg_value']/100:.2f}",
                    f"${at['avg_value']/100:.2f}",
                    f"${(at['avg_value']-mt['avg_value'])/100:+.2f}",
                ),
                (
                    "Avg value % of price",
                    f"{mt['avg_value_pct']:.1f}%",
                    f"{at['avg_value_pct']:.1f}%",
                    f"{at['avg_value_pct']-mt['avg_value_pct']:+.1f}pp",
                ),
                (
                    "Avg unique items / box",
                    f"{mt['avg_unique_items']:.1f}",
                    f"{at['avg_unique_items']:.1f}",
                    f"{at['avg_unique_items']-mt['avg_unique_items']:+.1f}",
                ),
                (
                    "Avg diversity score",
                    f"{mt['avg_diversity_score']:.3f}",
                    f"{at['avg_diversity_score']:.3f}",
                    f"{at['avg_diversity_score']-mt['avg_diversity_score']:+.3f}",
                ),
                (
                    "Avg bad dupes / box",
                    f"{mt['avg_bad_dupes']:.2f}",
                    f"{at['avg_bad_dupes']:.2f}",
                    f"{at['avg_bad_dupes']-mt['avg_bad_dupes']:+.2f}",
                ),
            ]
        elif mt:
            tier_rows = [
                ("Boxes", f"{mt['count']}", "n/a", ""),
                ("Avg value ($)", f"${mt['avg_value']/100:.2f}", "n/a", ""),
                ("Avg value % of price", f"{mt['avg_value_pct']:.1f}%", "n/a", ""),
            ]
        else:
            tier_rows = [
                ("Boxes", "n/a", f"{at['count']}", ""),
                ("Avg value ($)", "n/a", f"${at['avg_value']/100:.2f}", ""),
                ("Avg value % of price", "n/a", f"{at['avg_value_pct']:.1f}%", ""),
            ]

        for label, man_val, alg_val, diff in tier_rows:
            print(f"  {label:<30} {man_val:>12} {alg_val:>12} {diff:>10}")

    # --- Composite score breakdown ---
    print(f"\n{'='*70}")
    print(f"  COMPOSITE SCORE BREAKDOWN")
    print(f"{'='*70}")
    comp_header = f"  {'Component':<25} {'Manual':>10} {'Algorithm':>10}"
    print(comp_header)
    print(f"  {'-'*47}")
    for label, key in [
        ("Value penalty", "value_pen"),
        ("Same-item penalty", "si_pen"),
        ("Group concentration", "gc_pen"),
        ("Diversity penalty", "diversity_pen"),
        ("Max value share", "mvs_pen"),
        ("Size floor", "sf_pen"),
        ("Pref violation penalty", "pref_pen"),
    ]:
        if has_algo:
            print(f"  {label:<25} {-man_comp[key]:>+10.1f} {-alg_comp[key]:>+10.1f}")
        else:
            print(f"  {label:<25} {-man_comp[key]:>+10.1f} {'n/a':>10}")
    print(f"  {'-'*47}")
    if has_algo:
        print(f"  {'TOTAL SCORE':<25} {man_comp['score']:>10.1f} {alg_comp['score']:>10.1f}")
    else:
        print(f"  {'TOTAL SCORE':<25} {man_comp['score']:>10.1f} {'n/a':>10}")

    # --- Detail mode ---
    if args.detail:
        print_detail(algorithm, per_offer, all_algo)

    # --- CSV output ---
    if args.csv:
        alg_name = _effective_algorithm(algorithm)
        write_csv("manual", per_offer, "manual")
        write_csv(alg_name, per_offer, "algo")

    # --- Interpretation ---
    print(f"\n{'='*70}")
    print(f"  INTERPRETATION GUIDE")
    print(f"{'='*70}")
    print(f"  Composite score:    100 = perfect, >80 = good, >60 = acceptable, <60 = poor")
    print(f"  Value % of price:   114-117% = sweet spot (0 penalty)")
    print(f"                      110-120% = acceptable (small penalty)")
    print(f"                      <110% or >120% = heavier penalty (asymmetric: over-value softer)")
    print(f"  Diversity score:    1.0 = covers all available sub-cats, usages, colours, shapes")
    print(f"                      0.0 = no coverage; penalty = (1 - score) * 8.0")
    print(f"  Group-qty:          Excess items per fungible group above allowance")
    print(f"  Desirability:       Higher = more desirable items (from historical packing)")
    print(f"  Pref violations:    0 = all preferences respected (ideal)")
    print()


if __name__ == "__main__":
    main()
