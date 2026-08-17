"""
Historical data cleanup script.

Extracts the final allocation tab from each XLSX file, normalizes column
formats, classifies columns, and produces clean CSVs for algorithm validation.

Supports four archive tiers of data quality:
  A: Current standard layouts with IDs
  B: Earlier standard layouts with IDs
  C: Name-based layouts that need name matching
  D: Sparse rough formats with per-layout overrides
"""

import csv
import json
import logging
import re
import sys
from pathlib import Path

# When run as `python3 allocator/clean_history.py`, Python puts allocator/ on
# sys.path, causing allocator.py to shadow the allocator package. Fix by
# ensuring the project root is on the path.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from allocator.box_parser import classify_box
from allocator.config import (
    BOX_SIZE_OVERRIDES,
    BUFFER_IDENTIFIERS,
    CHARITY_IDENTIFIERS,
    SKIP_COLUMN_IDENTIFIERS,
    STOCK_IDENTIFIERS,
    SUM_IDENTIFIERS,
)

logger = logging.getLogger(__name__)

HISTORICAL_DIR = Path(__file__).parent.parent / "historical"
OLDER_DIR = HISTORICAL_DIR / "older"
CLEANED_DIR = Path(__file__).parent.parent / "cleaned"
CLEANED_LLM_DIR = Path(__file__).parent.parent / "cleaned_llm"
MAPPINGS_DIR = Path(__file__).parent.parent / "mappings"

# Offers to skip entirely (no usable mystery data)
# Note: "Week 6 Shopping List.xlsx" is also excluded — it has no offer_* prefix so
# discover_files() skips it automatically. It uses non-standard price tiers and is
# not useful for algorithm comparison or ML training.
# Offer 28: formula-driven workbook (pre-dates this project's structure); all
# per-box cells are Excel formulas with no cached values readable by openpyxl.
# Box values also don't match the 30/60/90 tier structure.
SKIP_OFFERS = {28, 44}

# Per-offer sheet overrides: {offer_id: sheet_name}
SHEET_OVERRIDES = {
    22: "Week 7 - 6-10-24 - offer_22_sho",  # shopping list has Box 1-8; "Mystery & Steve Box" is a totals summary
    30: "offer_30_shopping_list",             # Refunds sheet (last sheet) has Item col and gets selected erroneously
    32: "offer_32_shopping_list (1)",         # same — Refunds sheet has Item col
}

# Sheet names that should never be selected for allocation data
_NEVER_SELECT_SHEETS = {"refunds", "sales", "customers", "orders"}

# Preferred sheet names in priority order
_PREFERRED_SHEETS = [
    "Mystery Trello", "mystery trello",
    "Trello",
    "Mystery Data",
    "Mystery box Trello",
    "Mystery Boxes",
]

# Sheet names to try as fallback (calc tabs — have extra columns but usable)
_FALLBACK_SHEETS = [
    "Mystery Calc", "mystery calc", "Mystery",
    "Trello Mystery",
]


# ---------------------------------------------------------------------------
# Column classification
# ---------------------------------------------------------------------------

def classify_column(header: str | None) -> tuple[str, str | None]:
    """
    Classify a column header into a type.

    Returns:
        (type, size_tier) where type is one of:
        "id", "item", "merged", "standalone", "donation", "charity",
        "staff", "stock", "buffer", "sum", "skip"
        and size_tier is "small", "medium", "large", or None.
    """
    if header is None:
        return ("skip", None)

    header = str(header).strip()
    if not header:
        return ("skip", None)

    if header == "ID":
        return ("id", None)
    if header == "Item":
        return ("item", None)

    # Data columns that sometimes appear in the allocation tab
    if header in ("Price Ea", "Ov - Mys", "Qty Sold", "Required Buy", "Overage"):
        return ("skip", None)

    # Exact match against known special columns
    if header in CHARITY_IDENTIFIERS:
        return ("charity", None)
    if header in STOCK_IDENTIFIERS:
        return ("stock", None)
    if header in BUFFER_IDENTIFIERS:
        return ("buffer", None)
    if header in SUM_IDENTIFIERS:
        return ("sum", None)

    # Delegate box classification to box_parser (handles donation, staff,
    # merged/email, standalone, STANDALONE_NAME_TO_EMAIL, charity, ? stripping)
    _, size_tier, box_type = classify_box(header)
    if box_type == "donation":
        return ("donation", None)
    if box_type == "staff":
        return ("staff", None)
    if box_type == "merged":
        return ("merged", size_tier)
    return ("standalone", size_tier)


def classify_column_extended(header: str | None) -> tuple[str, str | None]:
    """
    Extended column classifier with box_parser for improved box detection.

    Handles older offers' naming conventions (Lge Charity, ?Sm DonorOrg, etc.)
    that the basic classify_column misses.
    """
    if header is None:
        return ("skip", None)

    h = str(header).strip()
    if not h:
        return ("skip", None)

    # Structural columns first
    if h == "ID":
        return ("id", None)
    if h in ("Item", "Name"):
        return ("item", None)

    # Extended skip patterns
    if h in SKIP_COLUMN_IDENTIFIERS:
        return ("skip", None)

    # Known special columns
    if h in STOCK_IDENTIFIERS:
        return ("stock", None)
    if h in BUFFER_IDENTIFIERS:
        return ("buffer", None)
    if h in SUM_IDENTIFIERS:
        return ("sum", None)

    # Numeric-only or empty-ish headers are skip columns
    try:
        float(h)
        return ("skip", None)
    except ValueError:
        pass

    # Use box_parser for box-type classification
    cleaned, size_tier, box_type = classify_box(h)

    if box_type == "donation":
        return ("donation", None)
    if box_type == "staff":
        return ("staff", None)
    if box_type == "merged":
        return ("merged", size_tier)

    # Check if this looks like a structural column we missed
    h_lower = h.lower()
    if any(kw in h_lower for kw in ("column", "actual", "total", "difference", "cost")):
        return ("skip", None)

    return ("standalone", size_tier)


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def extract_offer_id(filename: str) -> int | None:
    """Extract offer ID from filename like 'offer_104_shopping_list.xlsx'."""
    match = re.search(r"offer_(\d+)", filename)
    return int(match.group(1)) if match else None


def discover_files(*dirs: Path) -> dict[int, tuple[Path, str]]:
    """
    Scan directories for XLSX files and return best file per offer.

    Returns {offer_id: (filepath, source_dir_name)}.
    When duplicates exist across dirs, prefers files in earlier directories.
    """
    candidates: dict[int, list[tuple[Path, str]]] = {}

    for d in dirs:
        if not d.exists():
            continue
        dir_label = d.name if d.name != "historical" else "historical"
        if "older" in str(d):
            dir_label = "historical/older"

        for path in sorted(d.glob("offer_*_shopping_list*.xlsx")):
            offer_id = extract_offer_id(path.name)
            if offer_id is None:
                continue
            candidates.setdefault(offer_id, []).append((path, dir_label))

    result = {}
    for offer_id, paths in candidates.items():
        if len(paths) == 1:
            result[offer_id] = paths[0]
        else:
            # Prefer canonical name, then files in historical/ over older/
            def _score(entry):
                path, source = entry
                canonical = path.name == f"offer_{offer_id}_shopping_list.xlsx"
                in_main = "older" not in source
                has_final = "final" in path.name.lower()
                return (canonical, in_main, has_final, path.stat().st_size)

            best = max(paths, key=_score)
            result[offer_id] = best

    return result


# ---------------------------------------------------------------------------
# Sheet selection
# ---------------------------------------------------------------------------

def _sheet_has_column(ws, col_name: str, max_rows: int = 5) -> int | None:
    """Check if sheet has a column with given name in first max_rows. Returns row number or None."""
    for row_idx in range(1, min(max_rows + 1, (ws.max_row or 0) + 1)):
        for cell in ws.iter_cols(min_row=row_idx, max_row=row_idx):
            val = cell[0].value
            if val is not None and str(val).strip() == col_name:
                return row_idx
    return None


def select_allocation_sheet(wb, offer_id: int):
    """
    Select the best allocation sheet from a workbook.

    Returns (worksheet, sheet_name, header_row) or (None, None, None).
    header_row is 1-indexed.
    """
    sheets = wb.sheetnames

    # Check per-offer override
    if offer_id in SHEET_OVERRIDES:
        override = SHEET_OVERRIDES[offer_id]
        if override in sheets:
            ws = wb[override]
            hr = find_header_row(ws)
            return (ws, override, hr)

    # Try preferred sheet names (Trello tabs — cleanest data)
    for name in _PREFERRED_SHEETS:
        for sn in sheets:
            if sn.lower() == name.lower():
                ws = wb[sn]
                if ws.max_row and ws.max_row >= 2:
                    hr = find_header_row(ws)
                    return (ws, sn, hr)

    # Fall back to last sheet if it has ID or Item column (matches old behavior)
    # Skip sheets that are clearly not allocation data (refunds, sales, etc.)
    last_sheet = sheets[-1]
    if last_sheet.lower() not in _NEVER_SELECT_SHEETS:
        ws = wb[last_sheet]
        if ws.max_row and ws.max_row >= 2:
            hr = find_header_row(ws)
            headers = _read_header_row(ws, hr)
            header_strs = {str(h).strip() for h in headers if h}
            if "ID" in header_strs or "Item" in header_strs or "Name" in header_strs:
                return (ws, last_sheet, hr)

    # Try fallback sheets (Calc tabs — have extra columns but usable)
    for name in _FALLBACK_SHEETS:
        for sn in sheets:
            if sn.lower() == name.lower():
                ws = wb[sn]
                if ws.max_row and ws.max_row >= 2:
                    hr = find_header_row(ws)
                    return (ws, sn, hr)

    # Try the shopping list sheet if it has box-like columns
    ws = wb[sheets[0]]
    if ws.max_row and ws.max_row >= 2:
        hr = find_header_row(ws)
        headers = _read_header_row(ws, hr)
        # Check for box-like columns (Lge Charity, Med1, M Box 1, etc.)
        box_count = 0
        for h in headers:
            if h is None:
                continue
            _, _, btype = classify_box(str(h).strip())
            if btype in ("standalone", "merged", "donation"):
                box_count += 1
        if box_count >= 3:
            return (ws, sheets[0], hr)

    return (None, None, None)


def find_header_row(ws, max_rows: int = 5) -> int:
    """
    Find the row containing column headers (ID or Item).

    Returns 1-indexed row number. Defaults to 1 if not found.
    """
    for row_idx in range(1, min(max_rows + 1, (ws.max_row or 0) + 1)):
        row_vals = [cell.value for cell in list(ws.iter_rows(
            min_row=row_idx, max_row=row_idx))[0]]
        for val in row_vals:
            if val is not None:
                s = str(val).strip()
                if s in ("ID", "Item", "Name"):
                    return row_idx
    return 1


def _read_header_row(ws, row_num: int = 1) -> list:
    """Read a specific row as header values."""
    return [cell.value for cell in list(ws.iter_rows(
        min_row=row_num, max_row=row_num))[0]]


# ---------------------------------------------------------------------------
# Transposition detection
# ---------------------------------------------------------------------------

def _is_transposed(ws, header_row: int = 1) -> bool:
    """
    Check if a sheet has items as columns and boxes as rows (transposed).

    Heuristic: headers contain many long produce-like names (not emails),
    and column A rows 2+ have short box-like names.
    """
    headers = _read_header_row(ws, header_row)
    if len(headers) < 5:
        return False

    # If any header contains "@", this is a standard sheet with email columns
    for h in headers:
        if h and isinstance(h, str) and "@" in h:
            return False

    # If header has "ID", it's standard orientation
    for h in headers:
        if h and str(h).strip() == "ID":
            return False

    # Check if most header values (cols 1+) look like produce item names
    # (long strings without @ and not matching known column patterns)
    item_like = 0
    for h in headers[1:30]:  # sample first 30 columns
        if h and isinstance(h, str):
            s = h.strip()
            if len(s) > 15 and "@" not in s:
                item_like += 1

    if item_like < 5:
        return False

    # Check if column A values (rows after header) look like box names
    box_like = 0
    numeric_col_a = 0
    for row in ws.iter_rows(min_row=header_row + 1,
                            max_row=min(header_row + 5, ws.max_row or 0),
                            min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val and isinstance(val, str):
            stripped = str(val).strip()
            if len(stripped) < 30:
                box_like += 1
        elif val is not None and isinstance(val, (int, float)):
            numeric_col_a += 1

    if box_like >= 2:
        return True

    # If column A is numeric (row numbers), check column B for box names
    # (e.g., offer 53: "Number", "Box", then item names)
    if numeric_col_a >= 2 and len(headers) > 1:
        box_like_b = 0
        for row in ws.iter_rows(min_row=header_row + 1,
                                max_row=min(header_row + 5, ws.max_row or 0),
                                min_col=2, max_col=2, values_only=True):
            val = row[0]
            if val and isinstance(val, str):
                stripped = str(val).strip()
                if len(stripped) < 30:
                    box_like_b += 1
        if box_like_b >= 2:
            return True

    return False


def _read_transposed(ws, header_row: int = 1):
    """
    Read a transposed sheet (items as columns, boxes as rows).

    Returns (item_names, box_data) where:
        item_names = list of item names (from column headers)
        box_data = {box_header: {item_name: qty}}
    """
    headers = _read_header_row(ws, header_row)

    # Detect "Number"/"Box" pattern (offer 53): box names in col B, items from col C+
    h0 = str(headers[0]).strip().lower() if headers[0] else ""
    h1 = str(headers[1]).strip().lower() if len(headers) > 1 and headers[1] else ""
    if h0 in ("number", "#") and h1 in ("box", "name"):
        box_col = 1
        item_start = 2
    else:
        box_col = 0
        item_start = 1

    # Columns from item_start onward are item names
    item_names = []
    for h in headers[item_start:]:
        if h and str(h).strip():
            item_names.append(str(h).strip())
        else:
            break  # stop at first blank

    # Each subsequent row is a box
    box_data = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        box_name = row[box_col] if box_col < len(row) else None
        if box_name is None:
            continue
        box_name = str(box_name).strip()
        if not box_name:
            continue

        # Skip non-box rows (like "Pack order" metadata)
        if box_name.lower() in ("pack order", "total", "sum"):
            continue

        allocations = {}
        for i, item_name in enumerate(item_names):
            col = item_start + i
            val = row[col] if col < len(row) else None
            if val is not None:
                try:
                    qty = float(val)
                    if qty > 0:
                        allocations[item_name] = int(qty) if qty == int(qty) else qty
                except (ValueError, TypeError):
                    pass

        if allocations:
            box_data[box_name] = allocations

    return item_names, box_data


# ---------------------------------------------------------------------------
# Tier determination
# ---------------------------------------------------------------------------

def _determine_tier(offer_id: int, source_dir: str) -> str:
    """Determine data quality tier based on offer_id and source."""
    if offer_id >= 75:
        return "A"
    if offer_id >= 64:
        return "A"  # Tier A in historical/ dir (with quirks)
    if offer_id >= 55:
        return "B"
    if offer_id >= 45:
        return "C"
    return "D"


# ---------------------------------------------------------------------------
# Processing: with IDs (Tiers A, B)
# ---------------------------------------------------------------------------

def _classify_headers(headers, use_extended: bool = False):
    """
    Classify all column headers.

    Returns (classifications, id_col, item_col, mystery_cols, charity_cols, donation_cols).
    """
    classifier = classify_column_extended if use_extended else classify_column

    classifications = []
    id_col = None
    item_col = None
    mystery_cols = []
    charity_cols = []
    donation_cols = []

    for i, h in enumerate(headers):
        col_type, size_tier = classifier(h)
        classifications.append({
            "index": i,
            "header": str(h) if h else None,
            "type": col_type,
            "size_tier": size_tier,
        })

        if col_type == "id":
            id_col = i
        elif col_type == "item":
            item_col = i
        elif col_type in ("merged", "standalone"):
            mystery_cols.append((i, str(h), col_type, size_tier))
        elif col_type in ("charity", "donation"):
            # Donation boxes are charity recipients — include in charity CSV
            charity_cols.append((i, str(h)))
            if col_type == "donation":
                donation_cols.append((i, str(h)))

    return classifications, id_col, item_col, mystery_cols, charity_cols, donation_cols


def _read_data_rows(ws, header_row, id_col, mystery_cols, charity_cols):
    """Read allocation data from rows with an ID column."""
    rows = []
    charity_rows = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_id = row[id_col]
        if row_id is None:
            continue
        try:
            row_id = int(row_id)
        except (ValueError, TypeError):
            continue

        # Mystery box data
        row_data = {"id": row_id}
        for col_idx, col_header, col_type, _ in mystery_cols:
            val = row[col_idx] if col_idx < len(row) else None
            try:
                raw = float(val) if val and str(val).strip() else 0.0
            except (ValueError, TypeError):
                raw = 0.0
            row_data[col_header] = int(raw) if raw == int(raw) else raw
        rows.append(row_data)

        # Charity data
        if charity_cols:
            charity_data = {"id": row_id}
            for col_idx, col_header in charity_cols:
                val = row[col_idx] if col_idx < len(row) else None
                try:
                    raw = float(val) if val and str(val).strip() else 0.0
                except (ValueError, TypeError):
                    raw = 0.0
                charity_data[col_header] = int(raw) if raw == int(raw) else raw
            charity_rows.append(charity_data)

    return rows, charity_rows


def _dedup_rows(rows, charity_rows):
    """Merge rows sharing the same ID by summing quantities across box columns."""
    def _merge(row_list):
        by_id = {}
        for row in row_list:
            rid = row["id"]
            if rid in by_id:
                for k, v in row.items():
                    if k != "id":
                        by_id[rid][k] = by_id[rid].get(k, 0) + v
            else:
                by_id[rid] = dict(row)
        return list(by_id.values())
    return _merge(rows), _merge(charity_rows)


def _process_with_ids(ws, headers, header_row, offer_id, filepath,
                      sheet_name, source_dir, tier):
    """Process a sheet that has an ID column (Tiers A, B)."""
    use_extended = tier in ("B", "C", "D") or offer_id <= 74

    clsf, id_col, item_col, mystery_cols, charity_cols, donation_cols = \
        _classify_headers(headers, use_extended=use_extended)

    if id_col is None:
        return None

    rows, charity_rows = _read_data_rows(
        ws, header_row, id_col, mystery_cols, charity_cols)
    rows, charity_rows = _dedup_rows(rows, charity_rows)

    box_names = [h for _, h, _, _ in mystery_cols]
    box_sizes = {h: sz for _, h, _, sz in mystery_cols}

    metadata = {
        "offer_id": offer_id,
        "filename": filepath.name,
        "sheet_name": sheet_name,
        "total_items": len(rows),
        "box_count": len(mystery_cols),
        "box_names": box_names,
        "box_sizes": box_sizes,
        "box_types": {h: t for _, h, t, _ in mystery_cols},
        "charity_names": [h for _, h in charity_cols],
        "donation_names": [h for _, h in donation_cols],
        "classifications": clsf,
        "tier": tier,
        "source_dir": source_dir,
    }

    return {
        "metadata": metadata,
        "mystery_rows": rows,
        "charity_rows": charity_rows,
    }


# ---------------------------------------------------------------------------
# Processing: with names only (Tier C)
# ---------------------------------------------------------------------------

def _process_with_names(ws, headers, header_row, offer_id, filepath,
                        sheet_name, source_dir, tier):
    """Process a sheet with Item/Name column but no IDs (Tier C)."""
    clsf, id_col, item_col, mystery_cols, charity_cols, donation_cols = \
        _classify_headers(headers, use_extended=True)

    if item_col is None:
        logger.warning(f"Offer {offer_id}: no Item column found")
        return None

    if not mystery_cols:
        logger.warning(f"Offer {offer_id}: no box columns found")
        return None

    # Try to match item names to DB IDs
    try:
        from allocator.name_matcher import match_items
    except ImportError:
        logger.warning(f"Offer {offer_id}: name_matcher not available, skipping Tier C")
        return None

    # Collect item names and optional price column
    # Only include items that have a non-zero allocation to at least one
    # mystery box column — avoids matching full shopping list rows that
    # weren't allocated to any box (coffee, cheese, bakery, etc.)
    mystery_col_indices = {col_idx for col_idx, _, _, _ in mystery_cols}
    xlsx_names = []
    price_col = None
    for i, h in enumerate(headers):
        hs = str(h).strip() if h else ""
        if hs in ("JS Price Ea", "Price Ea"):
            price_col = i

    price_data = {}
    name_rows_raw = []  # (item_name, row_tuple)
    skipped_zero = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        item_name = row[item_col]
        if item_name is None:
            continue
        item_name = str(item_name).strip()
        if not item_name:
            continue

        # Skip summary/total rows (e.g. "TOTAL ITEMS" in offer 42)
        if item_name.upper() in {"TOTAL ITEMS", "TOTAL", "ITEMS TOTAL", "GRAND TOTAL"}:
            continue

        # Check if this item has any allocation to mystery boxes
        has_allocation = False
        for col_idx in mystery_col_indices:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                try:
                    if float(val) > 0:
                        has_allocation = True
                        break
                except (ValueError, TypeError):
                    pass

        if not has_allocation:
            skipped_zero += 1
            continue

        xlsx_names.append(item_name)
        name_rows_raw.append((item_name, row))

        if price_col is not None:
            pv = row[price_col]
            if pv is not None:
                try:
                    price_data[item_name] = float(pv)
                except (ValueError, TypeError):
                    pass

    if skipped_zero:
        logger.info(f"Offer {offer_id}: skipped {skipped_zero} items with no mystery box allocations")

    # Run name matching
    mappings = match_items(offer_id, xlsx_names,
                          price_column=price_data if price_data else None)

    # Build rows using matched IDs
    rows = []
    charity_rows = []
    matched = 0
    total = 0

    for item_name, row in name_rows_raw:
        total += 1
        if item_name not in mappings:
            continue
        matched += 1
        item_id = mappings[item_name]["id"]

        row_data = {"id": item_id}
        for col_idx, col_header, col_type, _ in mystery_cols:
            val = row[col_idx] if col_idx < len(row) else None
            try:
                raw = float(val) if val else 0.0
            except (ValueError, TypeError):
                raw = 0.0
            row_data[col_header] = int(raw) if raw == int(raw) else raw
        rows.append(row_data)

        if charity_cols:
            charity_data = {"id": item_id}
            for col_idx, col_header in charity_cols:
                val = row[col_idx] if col_idx < len(row) else None
                try:
                    raw = float(val) if val else 0.0
                except (ValueError, TypeError):
                    raw = 0.0
                charity_data[col_header] = int(raw) if raw == int(raw) else raw
            charity_rows.append(charity_data)

    match_quality = matched / total if total > 0 else 0.0
    rows, charity_rows = _dedup_rows(rows, charity_rows)

    box_names = [h for _, h, _, _ in mystery_cols]
    box_sizes = {h: sz for _, h, _, sz in mystery_cols}

    metadata = {
        "offer_id": offer_id,
        "filename": filepath.name,
        "sheet_name": sheet_name,
        "total_items": len(rows),
        "box_count": len(mystery_cols),
        "box_names": box_names,
        "box_sizes": box_sizes,
        "box_types": {h: t for _, h, t, _ in mystery_cols},
        "charity_names": [h for _, h in charity_cols],
        "donation_names": [h for _, h in donation_cols],
        "classifications": clsf,
        "tier": tier,
        "source_dir": source_dir,
        "name_match_quality": round(match_quality, 3),
    }

    return {
        "metadata": metadata,
        "mystery_rows": rows,
        "charity_rows": charity_rows,
    }


# ---------------------------------------------------------------------------
# Processing: transposed sheets
# ---------------------------------------------------------------------------

def _process_transposed(ws, header_row, offer_id, filepath, sheet_name,
                        source_dir, tier):
    """Process a transposed sheet (items as columns, boxes as rows)."""
    item_names, box_data = _read_transposed(ws, header_row)

    if not box_data:
        logger.warning(f"Offer {offer_id}: transposed sheet has no box data")
        return None

    # Filter to items that have at least one non-zero allocation
    allocated_items = set()
    for allocations in box_data.values():
        allocated_items.update(allocations.keys())
    items_to_match = [n for n in item_names if n in allocated_items]
    skipped_zero = len(item_names) - len(items_to_match)
    if skipped_zero:
        logger.info(f"Offer {offer_id}: skipped {skipped_zero} items with no allocations (transposed)")

    # Try to match item names to DB IDs
    try:
        from allocator.name_matcher import match_items
        mappings = match_items(offer_id, items_to_match)
    except ImportError:
        logger.warning(f"Offer {offer_id}: name_matcher not available for transposed")
        return None

    # Build rows: pivot from {box: {item: qty}} to {item_id: {box: qty}}
    item_to_id = {name: m["id"] for name, m in mappings.items()}
    all_item_ids = set()
    for allocations in box_data.values():
        for name in allocations:
            if name in item_to_id:
                all_item_ids.add(item_to_id[name])

    # Classify box names
    mystery_boxes = []
    charity_boxes = []
    donation_boxes = []
    for box_name in box_data:
        cleaned, size_tier, box_type = classify_box(box_name)
        if box_type in ("merged", "standalone"):
            mystery_boxes.append((box_name, box_type, size_tier))
        elif box_type in ("charity", "donation"):
            if box_type == "charity":
                charity_boxes.append(box_name)
            else:
                donation_boxes.append(box_name)

    # Donation boxes are charity recipients — include in charity CSV
    if not charity_boxes and donation_boxes:
        charity_boxes = donation_boxes

    if not mystery_boxes:
        logger.warning(f"Offer {offer_id}: no mystery boxes in transposed data")
        return None

    # Build row-oriented data (same format as _process_with_ids output)
    rows = []
    charity_rows = []
    for item_id in sorted(all_item_ids):
        # Find the item name for this ID
        item_name = None
        for name, mid in item_to_id.items():
            if mid == item_id:
                item_name = name
                break
        if item_name is None:
            continue

        row_data = {"id": item_id}
        for box_name, _, _ in mystery_boxes:
            qty = box_data.get(box_name, {}).get(item_name, 0)
            row_data[box_name] = int(qty) if qty == int(qty) else qty
        rows.append(row_data)

        if charity_boxes:
            charity_data = {"id": item_id}
            for box_name in charity_boxes:
                qty = box_data.get(box_name, {}).get(item_name, 0)
                charity_data[box_name] = int(qty) if qty == int(qty) else qty
            charity_rows.append(charity_data)

    box_names = [bn for bn, _, _ in mystery_boxes]
    box_sizes = {bn: sz for bn, _, sz in mystery_boxes}

    match_quality = len(item_to_id) / len(items_to_match) if items_to_match else 0.0
    rows, charity_rows = _dedup_rows(rows, charity_rows)

    metadata = {
        "offer_id": offer_id,
        "filename": filepath.name,
        "sheet_name": sheet_name,
        "total_items": len(rows),
        "box_count": len(mystery_boxes),
        "box_names": box_names,
        "box_sizes": box_sizes,
        "box_types": {bn: t for bn, t, _ in mystery_boxes},
        "charity_names": charity_boxes,
        "donation_names": donation_boxes,
        "classifications": [],
        "tier": tier,
        "source_dir": source_dir,
        "name_match_quality": round(match_quality, 3),
    }

    return {
        "metadata": metadata,
        "mystery_rows": rows,
        "charity_rows": charity_rows,
    }


# ---------------------------------------------------------------------------
# Processing: minimal (Tier D shopping list with embedded boxes)
# ---------------------------------------------------------------------------

def _process_minimal(ws, headers, header_row, offer_id, filepath,
                     sheet_name, source_dir, tier):
    """Best-effort processing for Tier D sheets with box columns in shopping list."""
    # Use extended classifier which handles Lge Charity, Med1, M Box 1, etc.
    clsf, id_col, item_col, mystery_cols, charity_cols, donation_cols = \
        _classify_headers(headers, use_extended=True)

    if not mystery_cols:
        logger.warning(f"Offer {offer_id}: no box columns found in minimal sheet")
        return None

    has_id = id_col is not None
    has_item = item_col is not None

    if has_id:
        return _process_with_ids(ws, headers, header_row, offer_id, filepath,
                                sheet_name, source_dir, tier)

    if has_item:
        return _process_with_names(ws, headers, header_row, offer_id, filepath,
                                  sheet_name, source_dir, tier)

    logger.warning(f"Offer {offer_id}: minimal sheet has no ID or Item column")
    return None


# ---------------------------------------------------------------------------
# Main processing
# ---------------------------------------------------------------------------

def process_file(filepath: Path, source_dir: str = "historical") -> dict | None:
    """
    Process a single XLSX file and extract allocation data.

    Returns metadata dict with column classifications and data,
    or None if the file can't be processed.
    """
    offer_id = extract_offer_id(filepath.name)
    if offer_id is None:
        print(f"  Skipping {filepath.name}: can't extract offer ID")
        return None

    if offer_id in SKIP_OFFERS:
        print(f"  Skipping offer {offer_id}: in SKIP_OFFERS")
        return None

    tier = _determine_tier(offer_id, source_dir)

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Select the best allocation sheet
    ws, sheet_name, header_row = select_allocation_sheet(wb, offer_id)
    if ws is None:
        print(f"  Skipping offer {offer_id}: no suitable sheet found")
        return None

    if ws.max_row is None or ws.max_row < 2:
        print(f"  Skipping offer {offer_id}: sheet '{sheet_name}' has < 2 rows")
        return None

    # Check for transposed orientation
    if _is_transposed(ws, header_row):
        print(f"  Offer {offer_id}: detected transposed sheet '{sheet_name}'")
        return _process_transposed(ws, header_row, offer_id, filepath,
                                   sheet_name, source_dir, tier)

    # Read headers
    headers = _read_header_row(ws, header_row)

    # Determine processing mode
    has_id = any(str(h).strip() == "ID" for h in headers if h)
    has_item = any(str(h).strip() in ("Item", "Name") for h in headers if h)

    if has_id:
        return _process_with_ids(ws, headers, header_row, offer_id, filepath,
                                sheet_name, source_dir, tier)
    elif has_item:
        return _process_with_names(ws, headers, header_row, offer_id, filepath,
                                  sheet_name, source_dir, tier)
    else:
        return _process_minimal(ws, headers, header_row, offer_id, filepath,
                               sheet_name, source_dir, tier)


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

def write_mystery_csv(offer_id: int, rows: list[dict], output_dir: Path) -> Path:
    """Write normalized mystery box allocation CSV."""
    if not rows:
        return None

    filepath = output_dir / f"offer_{offer_id}_mystery.csv"
    fieldnames = list(rows[0].keys())

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return filepath


def write_charity_csv(offer_id: int, rows: list[dict], output_dir: Path) -> Path:
    """Write normalized charity allocation CSV."""
    if not rows:
        return None

    # Only write if there's actual charity data (not just ID + all zeros)
    has_data = any(
        any(v for k, v in row.items() if k != "id")
        for row in rows
    )
    if not has_data:
        return None

    filepath = output_dir / f"offer_{offer_id}_charity.csv"
    fieldnames = list(rows[0].keys())

    with open(filepath, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return filepath


# ---------------------------------------------------------------------------
# LLM-based extraction (Tier C/D non-standard workbooks)
# ---------------------------------------------------------------------------

# Offers with XLSX files in historical/older/ that are candidates for LLM extraction.
# Excludes 44 (supplier purchasing list only — no mystery data).
LLM_TARGET_OFFERS = {
    22, 23, 24, 26, 28, 30, 32, 34, 36, 38, 40, 42,
    45, 46, 47, 48, 49, 50, 51, 52, 53, 54,
}


def _process_with_llm(
    offer_id: int,
    xlsx_path: Path,
    force: bool = False,
    method: str = "haiku-whole",
    workbook_info=None,
) -> dict | None:
    """
    Process a workbook using LLM extraction + name matching.

    Args:
        offer_id: The offer ID.
        xlsx_path: Path to the XLSX file.
        force: Ignore all caches and re-extract.
        method: Extraction strategy name (must be in benchmark_extraction.STRATEGY_RUNNERS).
        workbook_info: Pre-loaded WorkbookInfo (avoids redundant openpyxl loads when
            processing multiple offers in parallel). Loaded on-demand if None.

    Returns the same {metadata, mystery_rows, charity_rows} format as
    _process_with_ids/_process_with_names, or None on failure.
    """
    # Lazy import — benchmark_extraction imports from clean_history at module level,
    # so we must defer this import to avoid a circular dependency.
    from allocator import benchmark_extraction as _bx
    from allocator.name_matcher import match_items

    # 1. Check production cache: mappings/offer_N_llm_extraction_{method}.json
    prod_cache = MAPPINGS_DIR / f"offer_{offer_id}_llm_extraction_{method}.json"
    extraction = None
    if not force and prod_cache.exists():
        try:
            extraction = json.loads(prod_cache.read_text())
            logger.info(f"Offer {offer_id}: using production cache ({method})")
        except (json.JSONDecodeError, OSError):
            extraction = None

    # 2. Fall back to benchmark cache: benchmark_results/offer_N_{method}.json
    if extraction is None and not force:
        bench = _bx._load_cached(offer_id, method)
        if bench is not None and bench.success:
            extraction = dict(bench.raw_json)
            extraction.setdefault("sheet_used", bench.notes)
            extraction.setdefault("notes", bench.notes)
            logger.info(f"Offer {offer_id}: using benchmark cache ({method})")

    # 3. Run extraction via strategy runner
    if extraction is None:
        owns_workbook = False
        if workbook_info is None:
            workbook_info = _bx._load_workbook_info(offer_id, xlsx_path)
            owns_workbook = True
        try:
            result = _bx.STRATEGY_RUNNERS[method](offer_id, workbook_info)
            if not result.success:
                logger.warning(
                    f"Offer {offer_id}: {method} extraction failed: {result.error}"
                )
                return None
            extraction = dict(result.raw_json)
            extraction.setdefault("sheet_used", result.notes)
            extraction.setdefault("notes", result.notes)
            # Save to production cache
            try:
                MAPPINGS_DIR.mkdir(parents=True, exist_ok=True)
                prod_cache.write_text(json.dumps(extraction, indent=2))
            except OSError as e:
                logger.warning(f"Offer {offer_id}: could not save production cache: {e}")
        finally:
            if owns_workbook:
                workbook_info.cleanup()

    if extraction is None:
        return None

    boxes = extraction.get("boxes", [])
    items = extraction.get("items", [])
    charity_boxes = extraction.get("charity_boxes", [])
    charity_items = extraction.get("charity_items", [])

    if not boxes or not items:
        logger.warning(f"Offer {offer_id}: LLM extraction returned no boxes or items")
        return None

    # Collect item names for DB matching
    xlsx_names = [item["name"] for item in items]
    mappings = match_items(offer_id, xlsx_names)

    if not mappings:
        logger.warning(f"Offer {offer_id}: no items matched to DB")
        return None

    # Build box headers for CSV columns
    box_headers = [b["header"] for b in boxes]
    box_sizes = {}
    box_types = {}
    for b in boxes:
        header = b["header"]
        size = b.get("size", "unknown")
        box_sizes[header] = size if size != "unknown" else None
        box_types[header] = "standalone"

    # Build mystery rows
    mystery_rows = []
    for item in items:
        name = item["name"]
        if name not in mappings:
            continue
        item_id = mappings[name]["id"]
        row_data = {"id": item_id}
        for i, header in enumerate(box_headers):
            qty = item["allocations"][i] if i < len(item["allocations"]) else 0
            row_data[header] = int(qty) if isinstance(qty, float) and qty == int(qty) else qty
        mystery_rows.append(row_data)

    # Build charity rows
    charity_headers = [b["header"] for b in charity_boxes]
    charity_rows = []
    if charity_headers and charity_items:
        for item in charity_items:
            name = item["name"]
            if name not in mappings:
                continue
            item_id = mappings[name]["id"]
            row_data = {"id": item_id}
            for i, header in enumerate(charity_headers):
                qty = item["allocations"][i] if i < len(item["allocations"]) else 0
                row_data[header] = int(qty) if isinstance(qty, float) and qty == int(qty) else qty
            charity_rows.append(row_data)

    match_quality = len(mappings) / len(xlsx_names) if xlsx_names else 0.0
    mystery_rows, charity_rows = _dedup_rows(mystery_rows, charity_rows)

    metadata = {
        "offer_id": offer_id,
        "filename": xlsx_path.name,
        "sheet_name": extraction.get("sheet_used", "?"),
        "total_items": len(mystery_rows),
        "box_count": len(boxes),
        "box_names": box_headers,
        "box_sizes": box_sizes,
        "box_types": box_types,
        "charity_names": charity_headers,
        "donation_names": [],
        "classifications": [],
        "tier": _determine_tier(offer_id, "historical/older"),
        "source_dir": "historical/older",
        "name_match_quality": round(match_quality, 3),
        "extraction_method": f"llm-{method}",
        "llm_notes": extraction.get("notes", ""),
    }

    return {
        "metadata": metadata,
        "mystery_rows": mystery_rows,
        "charity_rows": charity_rows,
    }


def run_llm_extraction(
    offer_ids: set[int] | None = None,
    max_workers: int = 4,
    force: bool = False,
    methods: list[str] | None = None,
    progress_callback=None,
    cancel_check=None,
) -> dict:
    """
    Run LLM extraction on target offers with parallel workers and progress bar.

    Args:
        offer_ids: Specific offers to process (default: LLM_TARGET_OFFERS)
        max_workers: Concurrent LLM calls per method (default 4)
        force: Re-extract even if cached
        methods: Extraction method(s) to run (default: ["haiku-whole"]).
            Each method runs sequentially; offers within a method run in parallel.

    Returns summary dict keyed by method.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Lazy import to avoid circular dependency (benchmark_extraction imports clean_history)
    from allocator import benchmark_extraction as _bx

    if methods is None:
        methods = ["haiku-whole"]
    if offer_ids is None:
        offer_ids = LLM_TARGET_OFFERS

    # Discover target files once
    files = discover_files(HISTORICAL_DIR, OLDER_DIR)
    # Offer 28's Mystery Trello sheet uses text-encoded values readable by LLM
    # even though openpyxl can't parse its formula-driven cells.
    LLM_ALLOW = {28}
    targets = {
        oid: files[oid] for oid in sorted(offer_ids)
        if oid in files and (oid not in SKIP_OFFERS or oid in LLM_ALLOW)
    }

    if not targets:
        print("No target offers found.")
        return {}

    # Pre-load all workbooks once (avoids redundant openpyxl loads across methods)
    print(f"Pre-loading {len(targets)} workbooks...")
    workbooks: dict[int, object] = {}
    for oid, (filepath, _) in targets.items():
        try:
            workbooks[oid] = _bx._load_workbook_info(oid, filepath)
        except Exception as e:
            logger.warning(f"Offer {oid}: could not load workbook: {e}")

    overall_summary: dict[str, dict] = {}

    try:
        for method in methods:
            print(f"\n[{method}] {len(targets)} offers, {max_workers} workers")
            output_dir = CLEANED_LLM_DIR / method
            output_dir.mkdir(parents=True, exist_ok=True)

            summary: dict[str, dict] = {}
            errors: list[tuple[int, str]] = []

            def _process_one(offer_id: int, meth: str = method) -> tuple[int, dict | None, str | None]:
                filepath, _ = targets[offer_id]
                wb_info = workbooks.get(offer_id)
                try:
                    result = _process_with_llm(
                        offer_id, filepath,
                        force=force, method=meth, workbook_info=wb_info,
                    )
                    if result is None:
                        return (offer_id, None, "extraction returned None")
                    return (offer_id, result, None)
                except Exception as e:
                    return (offer_id, None, str(e))

            def _run_futures(executor, summary, errors, method):
                """Run futures and process results, using callback or Rich progress."""
                futures = {
                    executor.submit(_process_one, oid): oid for oid in targets
                }
                completed_count = 0
                total = len(targets)

                for future in as_completed(futures):
                    if cancel_check is not None and cancel_check():
                        executor.shutdown(wait=False, cancel_futures=True)
                        return

                    offer_id, result, error = future.result()
                    completed_count += 1

                    if error:
                        errors.append((offer_id, error))
                        if progress_callback is not None:
                            progress_callback(offer_id, completed_count, total)
                        continue

                    meta = result["metadata"]

                    # Write CSVs to method-specific output dir
                    write_mystery_csv(offer_id, result["mystery_rows"], output_dir)
                    write_charity_csv(offer_id, result["charity_rows"], output_dir)

                    summary[str(offer_id)] = meta

                    if progress_callback is not None:
                        progress_callback(offer_id, completed_count, total)

            if progress_callback is not None:
                # TUI mode: skip Rich Live display (conflicts with Textual's terminal)
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    _run_futures(executor, summary, errors, method)
            else:
                from rich.progress import BarColumn, MofNCompleteColumn, Progress, TextColumn

                with Progress(
                    TextColumn("[bold]{task.description}"),
                    BarColumn(),
                    MofNCompleteColumn(),
                    TextColumn("{task.fields[status]}"),
                ) as progress:
                    task = progress.add_task(
                        f"Extracting ({method})", total=len(targets), status=""
                    )

                    with ThreadPoolExecutor(max_workers=max_workers) as executor:
                        futures = {
                            executor.submit(_process_one, oid): oid for oid in targets
                        }

                        for future in as_completed(futures):
                            offer_id, result, error = future.result()

                            if error:
                                errors.append((offer_id, error))
                                progress.update(task, advance=1, status=f"offer {offer_id}: FAILED")
                                continue

                            meta = result["metadata"]

                            mystery_path = write_mystery_csv(offer_id, result["mystery_rows"], output_dir)
                            charity_path = write_charity_csv(offer_id, result["charity_rows"], output_dir)

                            status_parts = []
                            if mystery_path:
                                status_parts.append(
                                    f"offer {offer_id}: {meta['box_count']} boxes, "
                                    f"{meta['total_items']} items"
                                )
                            if charity_path:
                                status_parts.append("+ charity")

                            progress.update(
                                task, advance=1,
                                status=" ".join(status_parts) if status_parts else f"offer {offer_id}: no data",
                            )

                            summary[str(offer_id)] = meta

            # Print per-method summary
            print(f"\n[{method}] Completed: {len(summary)}/{len(targets)} offers")
            if errors:
                print(f"Errors ({len(errors)}):")
                for oid, err in sorted(errors):
                    print(f"  Offer {oid}: {err}")

            for oid in sorted(int(k) for k in summary):
                meta = summary[str(oid)]
                quality = meta.get("name_match_quality", 0)
                print(
                    f"  Offer {oid}: {meta['box_count']} boxes, "
                    f"{meta['total_items']} items, "
                    f"match {quality:.0%} — {meta.get('sheet_name', '?')}"
                )

            print(f"Output: {output_dir}/")
            overall_summary[method] = summary

        return overall_summary
    finally:
        for wb_info in workbooks.values():
            wb_info.cleanup()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def clean_all(historical_dir: Path = None, output_dir: Path = None,
              include_older: bool = True,
              only_offers: set[int] | None = None,
              progress_callback=None, cancel_check=None) -> dict:
    """
    Process all historical XLSX files and produce clean CSVs.

    Args:
        historical_dir: Primary directory (default: historical/)
        output_dir: Output directory for CSVs (default: cleaned/)
        include_older: Also scan historical/older/ (default: True)
        only_offers: If set, process only these offer IDs.

    Returns summary dict with metadata for all offers.
    """
    if historical_dir is None:
        historical_dir = HISTORICAL_DIR
    if output_dir is None:
        output_dir = CLEANED_DIR

    output_dir.mkdir(parents=True, exist_ok=True)

    # Discover files across directories
    dirs = [historical_dir]
    older_dir = historical_dir / "older"
    if include_older and older_dir.exists():
        dirs.append(older_dir)

    files = discover_files(*dirs)
    print(f"Found {len(files)} offers across {len(dirs)} directories")

    # Pre-filter to compute total for progress reporting
    target_ids = sorted(
        oid for oid in files.keys()
        if only_offers is None or oid in only_offers
    )
    total = len(target_ids)

    summary = {"offers": {}}

    for completed, offer_id in enumerate(target_ids):
        if cancel_check is not None and cancel_check():
            break
        if progress_callback is not None:
            progress_callback(offer_id, completed, total)
        filepath, source_dir = files[offer_id]
        print(f"\nProcessing offer {offer_id} ({source_dir})...")

        result = process_file(filepath, source_dir=source_dir)
        if result is None:
            continue

        meta = result["metadata"]

        # Write CSVs
        mystery_path = write_mystery_csv(offer_id, result["mystery_rows"], output_dir)
        charity_path = write_charity_csv(offer_id, result["charity_rows"], output_dir)

        if mystery_path:
            tier_label = f"[{meta.get('tier', '?')}]"
            print(f"  {tier_label} Mystery CSV: {mystery_path.name} "
                  f"({meta['box_count']} boxes, {meta['total_items']} items)")
            if "name_match_quality" in meta:
                print(f"  Name match quality: {meta['name_match_quality']:.0%}")
        if charity_path:
            print(f"  Charity CSV: {charity_path.name}")

        # Print column classification summary
        types_found = {}
        for c in meta.get("classifications", []):
            t = c["type"]
            types_found.setdefault(t, []).append(c["header"])
        for t, cols in sorted(types_found.items()):
            if t not in ("id", "item", "skip"):
                print(f"  {t}: {cols}")

        summary["offers"][str(offer_id)] = meta

    # Write summary JSON
    summary_path = output_dir / "summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"\nSummary written to {summary_path}")
    print(f"Processed {len(summary['offers'])} offers total")

    return summary


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Clean historical XLSX files")
    parser.add_argument("--no-older", action="store_true",
                       help="Skip historical/older/ directory")
    parser.add_argument("--only-offers", type=str, default=None,
                       help="Process only these offers, comma-separated with optional ranges "
                            "(e.g. 55-63 or 55,60,70-80)")
    parser.add_argument("--llm-extract", action="store_true",
                       help="Run LLM extraction on Tier C/D workbooks")
    parser.add_argument("--llm-method", type=str, default="haiku-whole",
                       help="LLM extraction method(s), comma-separated. "
                            "Default: haiku-whole. Options: sonnet-low, haiku-whole, "
                            "haiku-per-tab, sonnet-low-per-tab, two-stage-haiku, "
                            "two-stage-sonnet, two-stage-smart-haiku, two-stage-smart-sonnet")
    parser.add_argument("--force", action="store_true",
                       help="Force re-extraction (ignore cached results)")
    args = parser.parse_args()

    only_offers = None
    if args.only_offers:
        only_offers = set()
        for part in args.only_offers.split(","):
            part = part.strip()
            if "-" in part:
                lo, hi = part.split("-", 1)
                only_offers.update(range(int(lo), int(hi) + 1))
            else:
                only_offers.add(int(part))

    if args.llm_extract:
        from allocator import benchmark_extraction as _bx_validate
        methods = [m.strip() for m in args.llm_method.split(",")]
        invalid = [m for m in methods if m not in _bx_validate.STRATEGY_RUNNERS]
        if invalid:
            print(f"Unknown method(s): {invalid}")
            print(f"Available: {', '.join(_bx_validate.ALL_STRATEGIES)}")
            sys.exit(1)
        run_llm_extraction(force=args.force, methods=methods, offer_ids=only_offers)
    else:
        clean_all(include_older=not args.no_older, only_offers=only_offers)
